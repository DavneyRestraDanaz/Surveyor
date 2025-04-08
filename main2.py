import sys
import pandas as pd
from PyQt5.QtWidgets import (
    QApplication, QWidget, QPushButton, QFileDialog, QVBoxLayout, 
    QTableWidget, QTableWidgetItem, QLabel, QHBoxLayout, QLineEdit,
    QGridLayout, QGroupBox, QFormLayout, QHeaderView, QDialog, 
    QRadioButton, QDialogButtonBox, QCalendarWidget, QMessageBox,
    QComboBox, QScrollArea
)
from PyQt5.QtGui import (
    QFont, QTextDocument, QPageSize, QPageLayout,
    QPdfWriter, QPainter
)
from PyQt5.QtCore import (
    Qt, QEventLoop, QSizeF, QMarginsF, QUrl
)
from PyQt5.QtPrintSupport import QPrinter, QPrinterInfo, QPrintDialog, QPrintPreviewDialog
from PyQt5.QtWebEngineWidgets import QWebEngineView
import openpyxl
from openpyxl.writer.excel import ExcelWriter
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
import os
import shutil
import copy
from openpyxl.utils.cell import get_column_letter
import re
import base64

from openpyxl.styles import Font, Border, Side, Alignment, PatternFill

import random


class ExcelViewerApp(QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Aplikasi Laporan Psikologi")
        # Set window size to nearly fullscreen (90% of screen size)
        screen = QApplication.primaryScreen()
        screen_size = screen.availableGeometry()
        width = int(screen_size.width() * 0.9)
        height = int(screen_size.height() * 0.9)
        self.setGeometry(
            (screen_size.width() - width) // 2,  # Center horizontally
            (screen_size.height() - height) // 2, # Center vertically 
            width,
            height
        )
        # Updated columns based on the data from log
        self.columns = [
            "No", "No Tes", "Tgl Test", "TGL Lahir", "Nama PT", "JK", "SDR/SDRI", "Nama Peserta", 
            "PHQ", "Keterangan PHQ", "IQ ", "SE / Konkrit Praktis", "WA/ Verbal", " AN / Flexibilitas Pikir",
            "GE / Daya Abstraksi Verbal", "RA / Berpikir Praktis", "Unnamed: 16", "KLASIFIKASI",
            "N", "G", "A", "L", "P", "I", "T", "V", "S", "B", "O", "X", "C", "C (Coding)", "D", "R", "Z", "E", "K", "F", "W", 
            "NG", "CDR", "TV", "PI", "BS", "ZK",
            "Logika Berpikir 1", "Daya Analisa 3", "Kemampuan Verbal 2 dam 4", "Kemampuan Numerik 5", 
            "Sistematika Kerja/ C D R", "Orientasi Hasil/ N G", "Fleksibilitas/ T V", "Motivasi Berprestasi/ A", 
            "Kerjasama/ P I", "Keterampilan Interpersonal/ B S", "Stabilitas Emosi/ E PHQ", "Pegembangan Diri/ W", 
            "Mengelola Perubahan/ Z K",
            "Logika Berpikir 1.1", "Daya Analisa 3.1", "Kemampuan Verbal 2 dam 4.1", "Kemampuan Numerik 5.1", 
            "Sistematika Kerja/ C D R.1", "Orientasi Hasil/ N G.1", "Fleksibilitas/ T V.1", "Motivasi Berprestasi/ A.1", 
            "Kerjasama/ P I.1", "Keterampilan Interpersonal/ B S.1", "Stabilitas Emosi/ E PHQ.1", "Pegembangan Diri/ W.1", 
            "Mengelola Perubahan/ Z K.1"
        ]
        # Daftar kolom yang harus dihapus dari tampilan
        self.columns_to_hide = [
            'Unnamed: 13', 'Unnamed: 14', 'Intelegensi Umum.1', 'Daya Analisa/ AN.1', 
            'Kemampuan Verbal/WA GE.1', 'Kemampuan Numerik/ RA ZR.1', 'Daya Ingat/ME.1', 'Fleksibilitas',
            'Sistematika Kerja/ cd.1', 'Inisiatif/W.1', 'Stabilitas Emosi / E.1', 'Komunikasi / B O.1', 
            'Keterampilan Sosial / X S', 'Kerjasama'
        ]
        self.input_columns = self.columns.copy()
        self.initUI()
        self.excel_file_path = ""
        self.df = pd.DataFrame(columns=self.columns)

    def initUI(self):
        # Create scroll area
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        
        # Create container widget for scroll area
        container = QWidget()
        
        # Create main layout
        main_layout = QVBoxLayout(container)
        main_layout.setSpacing(20)
        main_layout.setContentsMargins(20, 20, 20, 20)
        
        # Set container as scroll area widget
        scroll.setWidget(container)
        
        # Create layout for the main window
        window_layout = QVBoxLayout(self)
        window_layout.addWidget(scroll)

        # File Selection Section
        file_group = QGroupBox("File Excel")
        file_group.setFont(QFont("Arial", 11, QFont.Bold))
        file_layout = QVBoxLayout()
        
        self.label = QLabel("Pilih file Excel (.xlsx)")
        self.label.setFont(QFont("Arial", 10))
        file_layout.addWidget(self.label)

        self.btn_select = QPushButton("Pilih File Excel") 
        self.btn_select.setFont(QFont("Arial", 10))
        self.btn_select.setFixedHeight(35)
        self.btn_select.clicked.connect(self.load_excel)
        # Remove automatic disable - now handled in load_excel() based on file selection
        file_layout.addWidget(self.btn_select)
        
        file_group.setLayout(file_layout)
        main_layout.addWidget(file_group)

        # Group for Personal Information
        personal_group = QGroupBox("Personal Information")
        personal_group.setFont(QFont("Arial", 11, QFont.Bold))
        personal_layout = QGridLayout()  # Changed to grid layout for better organization
        
        personal_fields = ["No", "No Tes", "Tgl Test", "TGL Lahir", "Nama PT", "JK", "SDR/SDRI", "Nama Peserta", "PHQ", "Keterangan PHQ"]
        self.personal_inputs = []
        for i, placeholder in enumerate(personal_fields):
            label = QLabel(placeholder + ":")
            label.setFont(QFont("Arial", 10))
            
            # Special handling for JK, TGL Lahir, dan Tgl Test
            if placeholder == "TGL Lahir" or placeholder == "Tgl Test":
                field = QPushButton("Pilih Tanggal")
                field.clicked.connect(lambda checked=False, placeholder=placeholder: self.show_calendar(placeholder))
            elif placeholder == "JK":
                field = QPushButton("Pilih Jenis Kelamin")
                field.clicked.connect(self.show_gender_dialog)
            elif placeholder == "SDR/SDRI" or placeholder == "Keterangan PHQ":
                # Read-only fields that will be auto-populated
                field = QLineEdit()
                field.setReadOnly(True)
                field.setStyleSheet("background-color: #f0f0f0;")  # Light gray background to indicate read-only
            else:
                field = QLineEdit()
                
            field.setFixedHeight(30)
            self.personal_inputs.append(field)
            row = i // 2  # Arrange fields in rows of 2
            col = (i % 2) * 2
            personal_layout.addWidget(label, row, col)
            personal_layout.addWidget(field, row, col + 1)
        
        personal_group.setLayout(personal_layout)
        main_layout.addWidget(personal_group)

        # Connect event handlers for auto-populating fields
        # PHQ field index is 8, Keterangan PHQ is 9
        self.personal_inputs[8].textChanged.connect(self.update_keterangan_phq)
        # JK field index is 5, SDR/SDRI is 6
        self.personal_inputs[5].clicked.connect(self.update_sdr_sdri)  # Will update after gender dialog closes

        # Group for IST
        ist_group = QGroupBox("IST")
        ist_group.setFont(QFont("Arial", 11, QFont.Bold))
        ist_layout = QGridLayout()  # Changed to grid layout
        
        # Sesuaikan field IST berdasarkan log
        ist_fields = ["IQ ", "SE / Konkrit Praktis", "WA/ Verbal", " AN / Flexibilitas Pikir", 
                      "GE / Daya Abstraksi Verbal", "RA / Berpikir Praktis", "Unnamed: 16"]
        self.ist_inputs = []
        for i, placeholder in enumerate(ist_fields):
            label = QLabel(placeholder + ":")
            label.setFont(QFont("Arial", 10))
            field = QLineEdit()
            field.setFixedHeight(30)
            
            # Make IQ and Unnamed:16 fields read-only
            if placeholder in ["IQ ", "Unnamed: 16"]:
                field.setReadOnly(True)
                field.setStyleSheet("background-color: #f0f0f0;")  # Light gray background to indicate read-only
            
            self.ist_inputs.append(field)
            row = i // 3  # Arrange fields in rows of 3
            col = (i % 3) * 2
            ist_layout.addWidget(label, row, col)
            ist_layout.addWidget(field, row, col + 1)
        
        ist_group.setLayout(ist_layout)
        main_layout.addWidget(ist_group)

        # Group for PAPIKOSTICK
        papikostick_group = QGroupBox("PAPIKOSTICK (Numeric)")
        papikostick_group.setFont(QFont("Arial", 11, QFont.Bold))
        papikostick_layout = QGridLayout()  # Changed to grid layout
        
        # Sesuaikan field PAPIKOSTICK berdasarkan log
        papikostick_fields = ["N", "G", "A", "L", "P", "I", "T", "V", "S", "B", "O", "X", 
                              "C", "D", "R", "Z", "E", "K", "F", "W"]  # W sekarang di posisi 19 (indeks 38 di Excel)
        self.papikostick_inputs = []
        for i, placeholder in enumerate(papikostick_fields):
            label = QLabel(placeholder + ":")
            label.setFont(QFont("Arial", 10))
            field = QLineEdit()
            field.setFixedHeight(30)
            self.papikostick_inputs.append(field)
            row = i // 5  # Arrange fields in rows of 5
            col = (i % 5) * 2
            papikostick_layout.addWidget(label, row, col)
            papikostick_layout.addWidget(field, row, col + 1)
        
        papikostick_group.setLayout(papikostick_layout)
        main_layout.addWidget(papikostick_group)

        # Add buttons for adding and updating data
        self.btn_add = QPushButton("Tambah Data")
        self.btn_add.setFont(QFont("Arial", 10))
        self.btn_add.setFixedHeight(35)
        self.btn_add.setFixedWidth(200)
        self.btn_add.clicked.connect(lambda: self.add_or_update_row("add"))
        self.btn_add.setEnabled(False)  # Initially disabled

        self.btn_edit = QPushButton("Edit Data")
        self.btn_edit.setFont(QFont("Arial", 10))
        self.btn_edit.setFixedHeight(35)
        self.btn_edit.setFixedWidth(200)
        self.btn_edit.clicked.connect(lambda: self.add_or_update_row("edit"))
        self.btn_edit.setEnabled(False)  # Initially disabled

        # Right-align the buttons in the layout
        button_container = QWidget()
        button_layout = QHBoxLayout(button_container)
        button_layout.addStretch()  # Push buttons to right
        button_layout.addWidget(self.btn_add)
        button_layout.addWidget(self.btn_edit)
        button_layout.setContentsMargins(0, 0, 20, 0)  # Add right margin
        main_layout.addWidget(button_container)

        # Enable buttons when file is loaded
        self.btn_select.clicked.connect(lambda: self.btn_add.setEnabled(True))
        self.btn_select.clicked.connect(lambda: self.btn_edit.setEnabled(True))

        # Add toggle buttons for each group
        toggle_layout = QHBoxLayout()
        toggle_personal = QPushButton("Personal Information")
        toggle_personal.setCheckable(True)
        toggle_personal.setChecked(False)  # Initially unchecked
        toggle_personal.setFont(QFont("Arial", 10))
        toggle_personal.setFixedHeight(35)
        toggle_personal.toggled.connect(lambda checked: personal_group.setVisible(checked))
        toggle_layout.addWidget(toggle_personal)
        
        toggle_ist = QPushButton("IST")
        toggle_ist.setCheckable(True)
        toggle_ist.setChecked(False)  # Initially unchecked
        toggle_ist.setFont(QFont("Arial", 10))
        toggle_ist.setFixedHeight(35)
        toggle_ist.toggled.connect(lambda checked: ist_group.setVisible(checked))
        toggle_layout.addWidget(toggle_ist)
        
        toggle_papikostick = QPushButton("PAPIKOSTICK")
        toggle_papikostick.setCheckable(True)
        toggle_papikostick.setChecked(False)  # Initially unchecked
        toggle_papikostick.setFont(QFont("Arial", 10))
        toggle_papikostick.setFixedHeight(35)
        toggle_papikostick.toggled.connect(lambda checked: papikostick_group.setVisible(checked))
        toggle_layout.addWidget(toggle_papikostick)
        
        # Initially hide all groups
        personal_group.setVisible(False)
        ist_group.setVisible(False)
        papikostick_group.setVisible(False)
        
        main_layout.addLayout(toggle_layout)

        # Search Section
        search_group = QGroupBox("Pencarian")
        search_group.setFont(QFont("Arial", 11, QFont.Bold))
        search_layout = QHBoxLayout()

        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("Cari data...")
        self.search_input.setFont(QFont("Arial", 10))
        self.search_input.setFixedHeight(30)
        self.search_input.textChanged.connect(self.search_table)
        search_layout.addWidget(self.search_input)

        self.search_column = QComboBox()
        self.search_column.setFont(QFont("Arial", 10))
        self.search_column.setFixedHeight(30)
        self.search_column.addItems(["Semua Kolom"] + self.columns)
        search_layout.addWidget(self.search_column)

        search_group.setLayout(search_layout)
        main_layout.addWidget(search_group)

        # Table Section
        table_group = QGroupBox("Data Hasil")
        table_group.setFont(QFont("Arial", 11, QFont.Bold))
        table_layout = QVBoxLayout()
        
        self.table = QTableWidget()
        self.table.setFont(QFont("Arial", 9))
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)  # Stretch columns to fit
        self.table.itemSelectionChanged.connect(self.populate_fields_from_selection)
        table_layout.addWidget(self.table)

        # Buttons below table
        button_layout = QHBoxLayout()
        
        self.btn_delete = QPushButton("Delete Row")
        self.btn_delete.setFont(QFont("Arial", 10))
        self.btn_delete.setFixedHeight(35)
        self.btn_delete.clicked.connect(self.delete_selected_row)
        button_layout.addWidget(self.btn_delete)

        # Add Preview PDF button
        self.btn_preview_pdf = QPushButton("Preview PDF")
        self.btn_preview_pdf.setFont(QFont("Arial", 10))
        self.btn_preview_pdf.setFixedHeight(35)
        self.btn_preview_pdf.clicked.connect(self.preview_pdf)
        button_layout.addWidget(self.btn_preview_pdf)

        self.btn_save_excel = QPushButton("Save to Excel")
        self.btn_save_excel.setFont(QFont("Arial", 10))
        self.btn_save_excel.setFixedHeight(35)
        self.btn_save_excel.clicked.connect(self.save_to_excel)
        button_layout.addWidget(self.btn_save_excel)
                
        table_layout.addLayout(button_layout)
        table_group.setLayout(table_layout)
        main_layout.addWidget(table_group)

        self.setLayout(main_layout)

    def search_table(self):
        # Check if Excel file has been loaded
        if not hasattr(self, 'excel_file_path') or not self.excel_file_path:
            QMessageBox.warning(self, "Warning", "Please load an Excel file first!")
            return

        search_text = self.search_input.text().lower().strip()
        selected_column = self.search_column.currentText()

        for row in range(self.table.rowCount()):
            row_visible = False

            if selected_column == "Semua Kolom":
                # Search in all columns
                for col in range(self.table.columnCount()):
                    item = self.table.item(row, col)
                    if item and search_text in item.text().lower():
                        row_visible = True
                        break
            else:
                # Search in selected column
                col_idx = self.table.horizontalHeader().logicalIndex(self.columns.index(selected_column))
                item = self.table.item(row, col_idx)
                if item and search_text in item.text().lower():
                    row_visible = True

            self.table.setRowHidden(row, not row_visible)

        # Debugging output
        print(f"Search text: '{search_text}' in column: '{selected_column}'")
        print(f"Total rows: {self.table.rowCount()}")
        for row in range(self.table.rowCount()):
            print(f"Row {row} visible: {not self.table.isRowHidden(row)}")

    def populate_fields_from_selection(self):
        selected_row = self.table.currentRow()
        if selected_row >= 0:
            # Definisikan papikostick_fields di sini
            papikostick_fields = ["N", "G", "A", "L", "P", "I", "T", "V", "S", "B", "O", "X", 
                                 "C", "D", "R", "Z", "E", "K", "F", "W"]  # W di posisi 19
            
            row_data = {}
            for col, column_name in enumerate(self.columns):
                item = self.table.item(selected_row, col)
                value = item.text() if item else ""
                # Ganti 'nan' dengan string kosong
                if value.lower() == 'nan':
                    value = ""
                row_data[column_name] = value

            # Debug prints
            print(f"Selected Row: {selected_row}")
            
            # Hanya tampilkan kolom yang digunakan (hapus log untuk kolom yang tidak dipakai)
            used_columns = [col for col in self.columns if col not in self.columns_to_hide]
            
            for key, value in row_data.items():
                if key in used_columns:
                    print(f"{key}: {value}")

            # Populate personal inputs
            for i, field in enumerate(self.personal_inputs):
                col_name = self.columns[i] if i < len(self.columns) else ""
                if col_name:
                    value = row_data.get(col_name, "")
                    # Skip kolom 6 (SDR/SDRI) dan 9 (Keterangan PHQ) karena akan diupdate otomatis
                    if i != 6 and i != 9:  
                        field.setText(value)
            
            # Manually update SDR/SDRI based on JK
            jk_value = self.personal_inputs[5].text()
            if jk_value == "P":
                self.personal_inputs[6].setText("Sdri.")
            elif jk_value == "L":
                self.personal_inputs[6].setText("Sdr.")
            else:
                self.personal_inputs[6].setText("")
                
            # Manually update Keterangan PHQ based on PHQ
            phq_text = self.personal_inputs[8].text().strip()
            if phq_text:
                try:
                    phq_value = float(phq_text)
                    keterangan_field = self.personal_inputs[9]
                    
                    # Apply formula: =@IFS(I4<5,"Tidak ada",I4<10,"Ringan",I4<15,"Sedang",I4<20,"Cukup Berat",I4<28,"Parah")
                    if phq_value < 5:
                        keterangan_field.setText("Tidak ada")
                    elif phq_value < 10:
                        keterangan_field.setText("Ringan")
                    elif phq_value < 15:
                        keterangan_field.setText("Sedang")
                    elif phq_value < 20:
                        keterangan_field.setText("Cukup Berat")
                    else:
                        keterangan_field.setText("Parah")
                except ValueError:
                    self.personal_inputs[9].setText("")
            else:
                self.personal_inputs[9].setText("")

            # Populate IST inputs
            ist_start_idx = len(self.personal_inputs)
            for i, field in enumerate(self.ist_inputs):
                col_idx = ist_start_idx + i
                if col_idx < len(self.columns):
                    value = row_data.get(self.columns[col_idx], "")
                    # Ganti 'nan' dengan string kosong
                    if value.lower() == 'nan':
                        value = ""
                    field.setText(value)
            
            # Debug untuk mapping PAPIKOSTICK
            print("DEBUG - Populate PAPIKOSTICK fields from selection:")
            
            # Telusuri setiap kolom PAPIKOSTICK untuk menemukan indeks yang tepat
            papiko_indices = {}
            for col_name in papikostick_fields:
                col_idx = self.get_column_index(col_name)
                if col_idx >= 0:
                    papiko_indices[col_name] = col_idx
                    
            # Telusuri setiap input field PAPIKOSTICK
            for field_idx, field_label in enumerate(papikostick_fields):
                if field_label in papiko_indices:
                    col_idx = papiko_indices[field_label]
                    value = self.get_cell_text(selected_row, col_idx)
                    
                    # Khusus untuk kolom W, ambil dari kolom 38
                    if field_label == "W":
                        value = self.get_cell_text(selected_row, 38)  # Ambil dari kolom 38
                    
                    # Ganti 'nan' dengan string kosong
                    if value.lower() == 'nan':
                        value = ""
                    
                    # Log debugging
                    print(f"DEBUG - Populate field PAPIKOSTICK {field_idx} ({field_label}) dengan nilai: '{value}'")
                    
                    # Skip kolom C (Coding) karena akan dihitung otomatis
                    if field_label == "C (Coding)":
                        print(f"DEBUG - Melewati field C (Coding) karena dihitung otomatis")
                        continue
                    
                    # Set nilai ke field input yang sesuai
                    if field_idx < len(self.papikostick_inputs):
                        self.papikostick_inputs[field_idx].setText(value)
                else:
                    print(f"DEBUG - Kolom {field_label} tidak ditemukan di tabel")
                    # Kosongkan kolom jika tidak ditemukan
                    if field_idx < len(self.papikostick_inputs):
                        self.papikostick_inputs[field_idx].setText("")

            # Debug untuk melihat posisi kolom W di Excel
            w_col_idx = 38  # Indeks kolom W di Excel
            print(f"DEBUG - Mencoba mengambil nilai W dari kolom {w_col_idx}")
            
            # Ambil nilai W langsung dari kolom 38
            w_value = self.get_cell_text(selected_row, w_col_idx)
            print(f"DEBUG - Nilai W yang diambil: '{w_value}'")
            
            # Simpan nilai W ke input field yang sesuai
            w_field_idx = papikostick_fields.index("W")  # Sekarang papikostick_fields sudah didefinisikan
            if w_field_idx < len(self.papikostick_inputs):
                self.papikostick_inputs[w_field_idx].setText(w_value)
                print(f"DEBUG - Nilai W diset ke input field: '{w_value}'")

    def load_excel(self):
        
        options = QFileDialog.Options()
        file_path, _ = QFileDialog.getOpenFileName(self, "Pilih File Excel", "", "Excel Files (*.xlsx);;All Files (*)", options=options)

        # Enable the button if user cancels file selection
        if not file_path:
            self.btn_select.setEnabled(True)
            return

        try:
            self.label.setText(f"File: {file_path}")
            self.excel_file_path = file_path
            # Keep button enabled so user can select another file
            self.btn_select.setEnabled(True)
            
            # Membaca seluruh sheet dalam file Excel dengan opsi keep_default_na=False
            # untuk mencegah nilai kosong menjadi NaN
            self.excel_data = pd.read_excel(file_path, sheet_name=None, keep_default_na=False)  # Baca semua sheet
            
            # Pastikan Sheet1 dan Sheet2 terbaca
            if "Sheet1" in self.excel_data:
                self.sheet1_data = self.excel_data["Sheet1"]
                # Ganti nilai NaN dengan string kosong
                self.sheet1_data = self.sheet1_data.fillna("").astype(str).replace("nan", "")
            else:
                print("Sheet1 tidak ditemukan!")
                self.btn_select.setEnabled(True)
                return

            if "Sheet2" in self.excel_data:
                self.sheet2_data = self.excel_data["Sheet2"]
                # Ganti nilai NaN dengan string kosong
                self.sheet2_data = self.sheet2_data.fillna("").astype(str).replace("nan", "")
            else:
                print("Sheet2 tidak ditemukan!")
                self.btn_select.setEnabled(True)
                return

            # Proses data setelah membaca
            self.process_excel(file_path)
            
        except Exception as e:
            print(f"Error loading file: {e}")
            self.btn_select.setEnabled(True)
            QMessageBox.critical(self, "Error", "Failed to load Excel file. Please try again.")

    def process_excel(self, file_path):
        try:
            # Baca semua sheet dalam file Excel
            sheets = pd.read_excel(file_path, sheet_name=None, engine='openpyxl', keep_default_na=False)

            # Pastikan Sheet1 dan Sheet2 ada
            if "Sheet1" in sheets:
                df_sheet1 = sheets["Sheet1"]
            else:
                print("Sheet1 tidak ditemukan!")
                df_sheet1 = None
                self.btn_select.setEnabled(True)
                return

            if "Sheet2" in sheets:
                df_sheet2 = sheets["Sheet2"]
            else:
                print("Sheet2 tidak ditemukan!")
                df_sheet2 = None
                self.btn_select.setEnabled(True)
                return

            # Proses Sheet1 (jika ada)
            if df_sheet1 is not None:
                # Tampilkan kolom asli untuk debugging
                original_cols = df_sheet1.columns.tolist()
                print("Original columns (Sheet1):", original_cols)

                # Mencari baris awal data berdasarkan keyword "No"
                start_row = None
                for idx, row in df_sheet1.iterrows():
                    if any(str(cell).strip().lower() == 'no' for cell in row):
                        start_row = idx
                        break

                if start_row is not None:
                    # Tambahkan parameter keep_default_na=False untuk mencegah nilai kosong menjadi NaN
                    df_sheet1 = pd.read_excel(file_path, sheet_name="Sheet1", engine='openpyxl', 
                                             skiprows=start_row+1, keep_default_na=False)
                    new_df = df_sheet1.copy()

                    # Ganti nilai 'nan' dengan string kosong untuk semua kolom
                    for col in new_df.columns:
                        new_df[col] = new_df[col].astype(str).replace('nan', '')
                        # Ganti nilai 'NaN' dengan string kosong
                        new_df[col] = new_df[col].replace('NaN', '')
                    
                    # Pastikan semua kolom PAPIKOSTICK ada dalam dataframe
                    papiko_columns = ["N", "G", "A", "L", "P", "I", "T", "V", "S", "B", "O", "X", "C", "D", "R", "Z", "E", "K", "F", "W"]
                    for col in papiko_columns:
                        if col not in new_df.columns:
                            new_df[col] = ""  # Tambahkan kolom jika tidak ada
                            print(f"Menambahkan kolom PAPIKOSTICK yang tidak ada: {col}")
                        elif new_df[col].isna().any() or (new_df[col] == "").any() or (new_df[col].str.lower() == "nan").any():
                            print(f"Mengisi nilai kosong/NaN pada kolom: {col}")
                            new_df[col] = new_df[col].replace(["", "nan", "NaN"], "0")

                    # Konversi kolom tertentu menjadi string
                    str_columns = ['No', 'No Tes', 'Tgl Test', 'TGL Lahir', 'Nama PT', 'JK', 'SDR/SDRI', 'Nama Peserta', 'PHQ', 'Keterangan PHQ']
                    for col in str_columns:
                        if col in new_df.columns:
                            new_df[col] = new_df[col].astype(str)
                            # Pastikan nilai kolom tidak 'nan'
                            new_df[col] = new_df[col].replace('nan', '')

                    # Konversi kolom angka ke numeric - sesuaikan dengan kolom yang dipakai
                    numeric_columns = ['IQ ', 'SE / Konkrit Praktis', 'WA/ Verbal', ' AN / Flexibilitas Pikir', 
                                       'GE / Daya Abstraksi Verbal', 'RA / Berpikir Praktis', 'Unnamed: 16']
                    for col in numeric_columns:
                        if col in new_df.columns:
                            # Simpan nilai asli tanpa konversi untuk kolom 'IQ ' dan 'Unnamed: 16'
                            if col == 'IQ ' or col == 'Unnamed: 16':
                                # Hanya ganti string kosong dan 'nan' dengan "" tetapi biarkan nilai lainnya
                                new_df[col] = new_df[col].replace(['nan', 'NaN'], '')
                                # Pastikan angka 0 string tidak diganti
                                new_df[col] = new_df[col].apply(lambda x: '' if x.strip() == '0' and col in ['IQ ', 'Unnamed: 16'] else x)
                            else:
                                # Untuk kolom lain, lakukan seperti biasa
                                new_df[col] = new_df[col].replace('', '0')
                                new_df[col] = pd.to_numeric(new_df[col], errors='coerce').fillna(0)
                        else:
                            # Tambahkan kolom yang tidak ada, dengan nilai kosong untuk IQ dan Unnamed:16
                            if col == 'IQ ' or col == 'Unnamed: 16':
                                new_df[col] = ''
                            else:
                                new_df[col] = 0

                    # Update SDR/SDRI berdasarkan JK
                    if 'JK' in new_df.columns and 'SDR/SDRI' in new_df.columns:
                        # Formula: =IF(F4="P","Sdri.","Sdr.")
                        def get_sdr_sdri(jk):
                            if jk == 'P':
                                return "Sdri."
                            elif jk == 'L':
                                return "Sdr."
                            else:
                                return ""
                        
                        new_df['SDR/SDRI'] = new_df['JK'].apply(get_sdr_sdri)
                        
                    # Update Keterangan PHQ berdasarkan PHQ
                    if 'PHQ' in new_df.columns and 'Keterangan PHQ' in new_df.columns:
                        def get_keterangan_phq(phq_str):
                            try:
                                if phq_str.strip() == '':
                                    return ''
                                phq = float(phq_str)
                                # Formula: =@IFS(I4<5,"Tidak ada",I4<10,"Ringan",I4<15,"Sedang",I4<20,"Cukup Berat",I4<28,"Parah")
                                if phq < 5:
                                    return "Tidak ada"
                                elif phq < 10:
                                    return "Ringan"
                                elif phq < 15:
                                    return "Sedang"
                                elif phq < 20:
                                    return "Cukup Berat"
                                else:
                                    return "Parah"
                            except (ValueError, AttributeError):
                                return ''
                                
                        new_df['Keterangan PHQ'] = new_df['PHQ'].apply(get_keterangan_phq)
                        
                    # Hitung IQ berdasarkan nilai-nilai IST
                    if all(col in new_df.columns for col in ["SE / Konkrit Praktis", "WA/ Verbal", " AN / Flexibilitas Pikir", "GE / Daya Abstraksi Verbal", "RA / Berpikir Praktis"]):
                        # Convert columns to numeric for calculation
                        numeric_cols = ["SE / Konkrit Praktis", "WA/ Verbal", " AN / Flexibilitas Pikir", "GE / Daya Abstraksi Verbal", "RA / Berpikir Praktis"]
                        for col in numeric_cols:
                            new_df[col] = pd.to_numeric(new_df[col], errors='coerce').fillna(0)
                        
                        # Formula: =SUM(L4:P4)/5
                        new_df['IQ '] = (new_df["SE / Konkrit Praktis"] + new_df["WA/ Verbal"] + new_df[" AN / Flexibilitas Pikir"] + 
                                         new_df["GE / Daya Abstraksi Verbal"] + new_df["RA / Berpikir Praktis"]) / 5
                    
                    # Hitung Unnamed: 16 = (AN + GE)/2
                    if all(col in new_df.columns for col in [" AN / Flexibilitas Pikir", "GE / Daya Abstraksi Verbal"]):
                        # Formula: =(M4+O4)/2
                        new_df['Unnamed: 16'] = (new_df[" AN / Flexibilitas Pikir"] + new_df["GE / Daya Abstraksi Verbal"]) / 2

                    # Konversi angka ke string untuk tampilan
                    for col in numeric_columns:
                        if col in new_df.columns and col not in ['IQ ', 'Unnamed: 16']:
                            new_df[col] = new_df[col].astype(str)

                    # Pastikan kolom W memiliki nilai
                    if "W" in new_df.columns:
                        # Jika ada nilai W di file Excel, gunakan nilai tersebut
                        # Pastikan nilai W tidak 'nan'
                        new_df["W"] = new_df["W"].replace('nan', '')
                    else:
                        # Jika tidak ada, buat kolom dengan nilai kosong
                        new_df["W"] = ""
                        print("Menambahkan kolom W dengan nilai kosong")

                    # Hitung KLASIFIKASI berdasarkan IQ
                    if "IQ " in new_df.columns:
                        # Gunakan konversi yang lebih aman untuk IQ
                        def safe_convert_iq(val):
                            try:
                                if val.strip() == '':
                                    return 0
                                return float(val)
                            except (ValueError, AttributeError):
                                return 0
                                
                        iq_vals = new_df["IQ "].apply(safe_convert_iq)
                        
                        def get_klasifikasi(iq):
                            # Formula: =@IFS(K4<79,"Rendah",K4<90,"Dibawah Rata-Rata",K4<110,"Rata-Rata",K4<120,"Diatas Rata-Rata",K4>119,"Superior")
                            if iq < 79:
                                return "Rendah"
                            elif 79 <= iq < 90:
                                return "Dibawah Rata-Rata"
                            elif 90 <= iq < 110:
                                return "Rata-Rata"
                            elif 110 <= iq < 120:
                                return "Diatas Rata-Rata"
                            else:
                                return "Superior"
                        
                        new_df["KLASIFIKASI"] = iq_vals.apply(get_klasifikasi)
                        print("Menghitung ulang nilai KLASIFIKASI berdasarkan IQ")

                    # Tambahkan kolom penting lainnya jika belum ada
                    important_columns = ["KLASIFIKASI"]
                    for col in important_columns:
                        if col not in new_df.columns:
                            print(f"Menambahkan kolom yang hilang: {col}")
                            new_df[col] = ""

                    # Tambahkan kolom psikogram (46-71)
                    psikogram_columns = [
                        "Logika Berpikir 1", "Daya Analisa 3", "Kemampuan Verbal 2 dam 4", 
                        "Kemampuan Numerik 5", "Sistematika Kerja/ C D R", "Orientasi Hasil/ N G",
                        "Fleksibilitas/ T V", "Motivasi Berprestasi/ A", "Kerjasama/ P I",
                        "Keterampilan Interpersonal/ B S", "Stabilitas Emosi/ E PHQ",
                        "Pegembangan Diri/ W", "Mengelola Perubahan/ Z K"
                    ]
                    
                    # Tambahkan kolom referensi (.1)
                    psikogram_ref_columns = [col + ".1" for col in psikogram_columns]
                    
                    # Pastikan semua kolom psikogram ada
                    for col in psikogram_columns + psikogram_ref_columns:
                        if col not in new_df.columns:
                            new_df[col] = ""
                            print(f"Menambahkan kolom psikogram yang tidak ada: {col}")

                    # Mapping untuk nama kolom input
                    input_column_mapping = {
                        "Sistematika Kerja/ C D R": "CDR",
                        "Orientasi Hasil/ N G": "NG",
                        "Fleksibilitas/ T V": "TV",
                        "Motivasi Berprestasi/ A": "A",  # Perhatikan ini hanya "A"
                        "Kerjasama/ P I": "PI",
                        "Keterampilan Interpersonal/ B S": "BS",
                        "Mengelola Perubahan/ Z K": "ZK"
                    }

                    # Mapping untuk kolom dasar
                    base_column_mapping = {
                        "Logika Berpikir 1": "SE / Konkrit Praktis",
                        "Daya Analisa 3": "WA/ Verbal",
                        "Kemampuan Verbal 2 dam 4": "GE / Daya Abstraksi Verbal",
                        "Kemampuan Numerik 5": " AN / Flexibilitas Pikir",
                        "Stabilitas Emosi/ E PHQ": "E",
                        "Pegembangan Diri/ W": "W"
                    }

                    # Mapping untuk kolom papikostick
                    papiko_column_mapping = {
                        "Sistematika Kerja/ C D R": "CDR",
                        "Orientasi Hasil/ N G": "NG",
                        "Fleksibilitas/ T V": "TV",
                        "Motivasi Berprestasi/ A": "A",
                        "Kerjasama/ P I": "PI",
                        "Keterampilan Interpersonal/ B S": "BS",
                        "Mengelola Perubahan/ Z K": "ZK"
                    }

                    # Debug: Print kolom yang ada
                    print("Kolom yang tersedia dalam DataFrame:")
                    print(new_df.columns.tolist())

                    # Hitung nilai untuk kolom psikogram (46-71)
                    for row_idx in range(len(new_df)):
                        try:
                            # Proses kolom-kolom dasar terlebih dahulu
                            for output_col, input_col in base_column_mapping.items():
                                try:
                                    print(f"Processing base column {output_col} using input column {input_col}")
                                    if input_col in new_df.columns:
                                        val = pd.to_numeric(new_df.iloc[row_idx][input_col], errors='coerce')
                                        print(f"Value for {input_col} = {val}")
                                        if pd.notna(val):
                                            # Untuk kolom E dan W, gunakan range PAPIKOSTICK
                                            if input_col in ["E", "W"]:
                                                if val < 2:
                                                    new_df.at[row_idx, output_col] = "R"
                                                elif val < 4:
                                                    new_df.at[row_idx, output_col] = "K"
                                                elif val < 6:
                                                    new_df.at[row_idx, output_col] = "C"
                                                elif val < 9:
                                                    new_df.at[row_idx, output_col] = "B"
                                                else:
                                                    new_df.at[row_idx, output_col] = "T"
                                            else:
                                                # Untuk kolom lain, gunakan range 80-140
                                                if val < 80:
                                                    new_df.at[row_idx, output_col] = "R"
                                                elif val < 100:
                                                    new_df.at[row_idx, output_col] = "K"
                                                elif val < 120:
                                                    new_df.at[row_idx, output_col] = "C"
                                                elif val < 140:
                                                    new_df.at[row_idx, output_col] = "B"
                                                else:
                                                    new_df.at[row_idx, output_col] = "T"
                                            print(f"Set {output_col} = {new_df.at[row_idx, output_col]}")
                                        else:
                                            print(f"Warning: NaN value for {input_col}, setting default value 'R'")
                                            new_df.at[row_idx, output_col] = "R"
                                    else:
                                        print(f"Warning: Column {input_col} not found in DataFrame")
                                except Exception as e:
                                    print(f"Error processing base column {output_col} for row {row_idx}: {e}")

                            # Proses kolom-kolom papikostick
                            for output_col, input_col in papiko_column_mapping.items():
                                try:
                                    print(f"Processing papiko column {output_col} using input column {input_col}")
                                    if input_col in new_df.columns:
                                        # Coba ambil nilai dari kolom kombinasi terlebih dahulu
                                        if input_col in ["CDR", "NG", "TV", "PI", "BS", "ZK"]:
                                            # Ambil nilai dari kolom-kolom yang membentuk kombinasi
                                            if input_col == "CDR":
                                                c_val = pd.to_numeric(new_df.iloc[row_idx]["C"], errors='coerce')
                                                d_val = pd.to_numeric(new_df.iloc[row_idx]["D"], errors='coerce')
                                                r_val = pd.to_numeric(new_df.iloc[row_idx]["R"], errors='coerce')
                                                if pd.notna(c_val) and pd.notna(d_val) and pd.notna(r_val):
                                                    val = (c_val + d_val + r_val) / 3
                                                else:
                                                    val = float('nan')
                                            elif input_col == "NG":
                                                n_val = pd.to_numeric(new_df.iloc[row_idx]["N"], errors='coerce')
                                                g_val = pd.to_numeric(new_df.iloc[row_idx]["G"], errors='coerce')
                                                if pd.notna(n_val) and pd.notna(g_val):
                                                    val = (n_val + g_val) / 2
                                                else:
                                                    val = float('nan')
                                            elif input_col == "TV":
                                                t_val = pd.to_numeric(new_df.iloc[row_idx]["T"], errors='coerce')
                                                v_val = pd.to_numeric(new_df.iloc[row_idx]["V"], errors='coerce')
                                                if pd.notna(t_val) and pd.notna(v_val):
                                                    val = (t_val + v_val) / 2
                                                else:
                                                    val = float('nan')
                                            elif input_col == "PI":
                                                p_val = pd.to_numeric(new_df.iloc[row_idx]["P"], errors='coerce')
                                                i_val = pd.to_numeric(new_df.iloc[row_idx]["I"], errors='coerce')
                                                if pd.notna(p_val) and pd.notna(i_val):
                                                    val = (p_val + i_val) / 2
                                                else:
                                                    val = float('nan')
                                            elif input_col == "BS":
                                                b_val = pd.to_numeric(new_df.iloc[row_idx]["B"], errors='coerce')
                                                s_val = pd.to_numeric(new_df.iloc[row_idx]["S"], errors='coerce')
                                                if pd.notna(b_val) and pd.notna(s_val):
                                                    val = (b_val + s_val) / 2
                                                else:
                                                    val = float('nan')
                                            elif input_col == "ZK":
                                                z_val = pd.to_numeric(new_df.iloc[row_idx]["Z"], errors='coerce')
                                                k_val = pd.to_numeric(new_df.iloc[row_idx]["K"], errors='coerce')
                                                if pd.notna(z_val) and pd.notna(k_val):
                                                    val = (z_val + k_val) / 2
                                                else:
                                                    val = float('nan')
                                        else:
                                            val = pd.to_numeric(new_df.iloc[row_idx][input_col], errors='coerce')
                                        
                                        print(f"Value for {input_col} = {val}")
                                        if pd.notna(val):
                                            if val < 2:
                                                new_df.at[row_idx, output_col] = "R"
                                            elif val < 4:
                                                new_df.at[row_idx, output_col] = "K"
                                            elif val < 6:
                                                new_df.at[row_idx, output_col] = "C"
                                            elif val < 9:
                                                new_df.at[row_idx, output_col] = "B"
                                            else:
                                                new_df.at[row_idx, output_col] = "T"
                                            print(f"Set {output_col} = {new_df.at[row_idx, output_col]}")
                                        else:
                                            print(f"Warning: NaN value for {input_col}, setting default value 'R'")
                                            new_df.at[row_idx, output_col] = "R"
                                    else:
                                        print(f"Warning: Column {input_col} not found in DataFrame")
                                except Exception as e:
                                    print(f"Error processing papiko column {output_col} for row {row_idx}: {e}")

                            # Untuk kolom .1 (59-71), gunakan get_sheet3_reference
                            for base_col in psikogram_columns:
                                try:
                                    ref_col = base_col + ".1"
                                    base_val = new_df.at[row_idx, base_col]
                                    print(f"Processing reference for {base_col} with value {base_val}")
                                    if pd.notna(base_val) and base_val.strip() != "":
                                        ref_val = self.get_sheet3_reference(base_col, base_val)
                                        new_df.at[row_idx, ref_col] = ref_val
                                        print(f"Set {ref_col} = {ref_val}")
                                    else:
                                        print(f"Warning: Empty or NaN value for {base_col}, getting reference for default 'R'")
                                        ref_val = self.get_sheet3_reference(base_col, "R")
                                        new_df.at[row_idx, ref_col] = ref_val
                                except Exception as e:
                                    print(f"Error processing reference for {base_col}: {e}")

                        except Exception as e:
                            print(f"Error pada baris {row_idx}: {e}")
                            continue

                    # Pastikan tidak ada nilai NaN yang tersisa
                    self.df_sheet1 = new_df.fillna("")
                    self.columns = list(new_df.columns)
                    self.show_table(self.df_sheet1)
                    
                    # Recalculate values for all rows
                    for row in range(self.table.rowCount()):
                        self.recalculate_values(row)
                        
                else:
                    print("Could not find the start of data in Sheet1")
                    self.btn_select.setEnabled(True)
                    return

            # Proses Sheet2 (jika ada)
            if df_sheet2 is not None:
                print("Original columns (Sheet2):", df_sheet2.columns.tolist())

                # Simpan data Sheet2 tanpa pemrosesan tambahan
                self.df_sheet2 = df_sheet2.fillna("")
        
        except Exception as e:
            print(f"Error loading Excel file: {e}")
            import traceback
            traceback.print_exc()
            self.btn_select.setEnabled(True)
            QMessageBox.critical(self, "Error", "Failed to process Excel file. Please try again.")

    def show_table(self, df):
        try:
            # Pastikan kolom-kolom penting ada
            important_columns = ["KLASIFIKASI"]
            df_with_columns = df.copy()
            
            # Tambahkan kolom yang hilang
            for col in important_columns:
                if col not in df_with_columns.columns:
                    print(f"Menambahkan kolom yang hilang: {col}")
                    df_with_columns[col] = ""
            
            # Setup table untuk menampilkan data
            self.table.setRowCount(df_with_columns.shape[0])
            self.table.setColumnCount(len(df_with_columns.columns))
            self.table.setHorizontalHeaderLabels(df_with_columns.columns)
            
            # Populate table with data
            for i, row in df_with_columns.iterrows():
                for j, (col_name, val) in enumerate(row.items()):
                    item = QTableWidgetItem(str(val))
                    self.table.setItem(i, j, item)
                    # Debug untuk kolom W
                    if col_name == "W":
                        print(f"DEBUG - Show_table: Mengatur nilai W='{val}' ke tabel di baris {i}, kolom {j}")
            
            # Sembunyikan kolom-kolom yang tidak dipakai
            for col_name in self.columns_to_hide:
                if col_name in df_with_columns.columns:
                    col_idx = df_with_columns.columns.get_loc(col_name)
                    self.table.hideColumn(col_idx)
                    print(f"Menyembunyikan kolom yang tidak dipakai: {col_name} (indeks {col_idx})")
            
            # Resize columns to fit content
            self.table.resizeColumnsToContents()
            
            # Perbarui daftar kolom
            self.columns = list(df_with_columns.columns)
            
            print("Berhasil menampilkan tabel")
        except Exception as e:
            print(f"Error saat menampilkan tabel: {e}")
            import traceback
            traceback.print_exc()

    def show_calendar(self, field_type):
        calendar = QCalendarWidget(self)
        calendar.setWindowFlags(Qt.Popup)
        
        if field_type == "TGL Lahir":
            # For TGL Lahir, use the 3rd input field
            target_field = self.personal_inputs[3]
            calendar.clicked.connect(lambda date: self.set_date(date, "TGL Lahir"))
        elif field_type == "Tgl Test":
            # For Tgl Test, use the 2nd input field
            target_field = self.personal_inputs[2]
            calendar.clicked.connect(lambda date: self.set_date(date, "Tgl Test"))
        
        # Position calendar below the button
        pos = target_field.mapToGlobal(target_field.rect().bottomLeft())
        calendar.move(pos)
        calendar.show()

    def set_date(self, date, field_type):
        # Format date as dd/MM/yyyy
        formatted_date = date.toString("dd/MM/yyyy")
        
        if field_type == "TGL Lahir":
            # Set TGL Lahir field (index 3)
            self.personal_inputs[3].setText(formatted_date)
        elif field_type == "Tgl Test":
            # Set Tgl Test field (index 2)
            self.personal_inputs[2].setText(formatted_date)
            
    def update_sdr_sdri(self):
        # This will be called after gender dialog closes
        jk_value = self.personal_inputs[5].text()
        sdr_sdri_field = self.personal_inputs[6]
        
        # Formula: =IF(F4="P","Sdri.","Sdr.")
        if jk_value == "P":
            sdr_sdri_field.setText("Sdri.")
        elif jk_value == "L":
            sdr_sdri_field.setText("Sdr.")
        else:
            sdr_sdri_field.setText("")
            
    def update_keterangan_phq(self):
        try:
            # Get PHQ value
            phq_text = self.personal_inputs[8].text().strip()
            keterangan_field = self.personal_inputs[9]
            
            if not phq_text:
                keterangan_field.setText("")
                return
                
            phq_value = float(phq_text)
            
            # Apply formula: =@IFS(I4<5,"Tidak ada",I4<10,"Ringan",I4<15,"Sedang",I4<20,"Cukup Berat",I4<28,"Parah")
            if phq_value < 5:
                keterangan_field.setText("Tidak ada")
            elif phq_value < 10:
                keterangan_field.setText("Ringan")
            elif phq_value < 15:
                keterangan_field.setText("Sedang")
            elif phq_value < 20:
                keterangan_field.setText("Cukup Berat")
            else:
                keterangan_field.setText("Parah")
        except ValueError:
            # Handle case where PHQ is not a valid number
            self.personal_inputs[9].setText("")

    def show_gender_dialog(self):
        from PyQt5.QtWidgets import QDialog, QVBoxLayout, QRadioButton, QDialogButtonBox
        
        dialog = QDialog(self)
        dialog.setWindowTitle("Pilih Jenis Kelamin")
        layout = QVBoxLayout()
        
        radio_l = QRadioButton("L (Laki-laki)")
        radio_p = QRadioButton("P (Perempuan)")
        
        layout.addWidget(radio_l)
        layout.addWidget(radio_p)
        
        buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        buttons.accepted.connect(dialog.accept)
        buttons.rejected.connect(dialog.reject)
        
        layout.addWidget(buttons)
        dialog.setLayout(layout)
        
        if dialog.exec_() == QDialog.Accepted:
            if radio_l.isChecked():
                self.personal_inputs[5].setText("L")
            elif radio_p.isChecked():
                self.personal_inputs[5].setText("P")
            
            # Update SDR/SDRI field based on gender
            self.update_sdr_sdri()

    def calculate_psikogram_values(self, ist_values, papikostick_values):
        """Menghitung nilai-nilai psikogram berdasarkan rumus yang diberikan"""
        psikogram = {}
        
        # Fungsi helper untuk menentukan nilai berdasarkan range IST
        def get_ist_value(value):
            try:
                value = float(value)
                if value < 80:
                    return "R"
                elif value < 100:
                    return "K"
                elif value < 120:
                    return "C"
                elif value < 140:
                    return "B"
                else:
                    return "T"
            except (ValueError, TypeError):
                return ""

        # Fungsi helper untuk menentukan nilai berdasarkan range PAPIKOSTICK
        def get_papiko_value(value):
            try:
                value = float(value)
                if value < 2:
                    return "R"
                elif value < 4:
                    return "K"
                elif value < 6:
                    return "C"
                elif value < 9:
                    return "B"
                elif value == 9:
                    return "T"
                else:
                    return ""
            except (ValueError, TypeError):
                return ""

        # Logika Berpikir 1 (L)
        psikogram['Logika Berpikir 1'] = get_ist_value(ist_values.get("SE / Konkrit Praktis", ""))

        # Daya Analisa 3 (M)
        psikogram['Daya Analisa 3'] = get_ist_value(ist_values.get("WA/ Verbal", ""))

        # Kemampuan Verbal 2 dam 4 (O)
        psikogram['Kemampuan Verbal 2 dam 4'] = get_ist_value(ist_values.get("GE / Daya Abstraksi Verbal", ""))

        # Kemampuan Numerik 5 (N)
        psikogram['Kemampuan Numerik 5'] = get_ist_value(ist_values.get(" AN / Flexibilitas Pikir", ""))

        # Sistematika Kerja/ C D R
        psikogram['Sistematika Kerja/ C D R'] = get_papiko_value(papikostick_values.get("CDR", ""))

        # Orientasi Hasil/ N G
        psikogram['Orientasi Hasil/ N G'] = get_papiko_value(papikostick_values.get("NG", ""))

        # Fleksibilitas/ T V
        psikogram['Fleksibilitas/ T V'] = get_papiko_value(papikostick_values.get("TV", ""))

        # Motivasi Berprestasi/ A
        psikogram['Motivasi Berprestasi/ A'] = get_papiko_value(papikostick_values.get("A", ""))

        # Kerjasama/ P I
        psikogram['Kerjasama/ P I'] = get_papiko_value(papikostick_values.get("PI", ""))

        # Keterampilan Interpersonal/ B S
        psikogram['Keterampilan Interpersonal/ B S'] = get_papiko_value(papikostick_values.get("BS", ""))

        # Stabilitas Emosi/ E PHQ
        psikogram['Stabilitas Emosi/ E PHQ'] = get_papiko_value(papikostick_values.get("E", ""))

        # Pegembangan Diri/ W
        psikogram['Pegembangan Diri/ W'] = get_papiko_value(papikostick_values.get("W", ""))

        # Mengelola Perubahan/ Z K
        psikogram['Mengelola Perubahan/ Z K'] = get_papiko_value(papikostick_values.get("ZK", ""))

        # Tambahkan kolom .1 dengan nilai yang sama
        for base_key in list(psikogram.keys()):
            if not base_key.endswith('.1'):
                psikogram[base_key + '.1'] = psikogram[base_key]

        return psikogram
    def add_or_update_row(self, mode="add"):
        try:
            # Collect values from all three input groups
            values = {}
            
            # For personal fields (kolom 0-9)
            personal_fields = ["No", "No Tes", "Tgl Test", "TGL Lahir", "Nama PT", "JK", "SDR/SDRI", "Nama Peserta", "PHQ", "Keterangan PHQ"]
            for i, field in enumerate(self.personal_inputs):
                field_name = personal_fields[i]
                
                # Penanganan khusus untuk SDR/SDRI
                if field_name == "SDR/SDRI":
                    jk_value = self.personal_inputs[5].text()
                    if jk_value == "P":
                        values[field_name] = "Sdri."
                    elif jk_value == "L":
                        values[field_name] = "Sdr."
                    else:
                        values[field_name] = ""
                        
                # Penanganan khusus untuk Keterangan PHQ
                elif field_name == "Keterangan PHQ":
                    phq_text = self.personal_inputs[8].text().strip()
                    if phq_text:
                        try:
                            phq_value = float(phq_text)
                            if phq_value < 5:
                                values[field_name] = "Tidak ada"
                            elif phq_value < 10:
                                values[field_name] = "Ringan"
                            elif phq_value < 15:
                                values[field_name] = "Sedang"
                            elif phq_value < 20:
                                values[field_name] = "Cukup Berat"
                            else:
                                values[field_name] = "Parah"
                        except ValueError:
                            values[field_name] = ""
                    else:
                        values[field_name] = ""
                else:
                    values[field_name] = field.text()

            # Add IST values (kolom 10-16)
            ist_fields = ["IQ ", "SE / Konkrit Praktis", "WA/ Verbal", " AN / Flexibilitas Pikir", 
                         "GE / Daya Abstraksi Verbal", "RA / Berpikir Praktis", "Unnamed: 16"]
            for i, field in enumerate(self.ist_inputs):
                field_name = ist_fields[i]
                value = field.text().strip()
                try:
                    if value:
                        float_value = float(value)
                        values[field_name] = float_value
                    else:
                        values[field_name] = ""
                except ValueError:
                    values[field_name] = value

            # Calculate IQ and Unnamed: 16
            if all(values.get(field, "") != "" for field in ["SE / Konkrit Praktis", "WA/ Verbal", " AN / Flexibilitas Pikir", 
                                                            "GE / Daya Abstraksi Verbal", "RA / Berpikir Praktis"]):
                # Calculate IQ = SUM(L:P)/5
                iq_values = [
                    float(values.get("SE / Konkrit Praktis", 0)),
                    float(values.get("WA/ Verbal", 0)),
                    float(values.get(" AN / Flexibilitas Pikir", 0)),
                    float(values.get("GE / Daya Abstraksi Verbal", 0)),
                    float(values.get("RA / Berpikir Praktis", 0))
                ]
                values["IQ "] = sum(iq_values) / 5

                # Calculate Unnamed: 16 = (WA + GE)/2
                values["Unnamed: 16"] = (float(values.get("WA/ Verbal", 0)) + 
                                       float(values.get("GE / Daya Abstraksi Verbal", 0))) / 2

                # Calculate KLASIFIKASI based on IQ
                iq_value = values["IQ "]
                if iq_value < 79:
                    values["KLASIFIKASI"] = "Rendah"
                elif iq_value < 90:
                    values["KLASIFIKASI"] = "Dibawah Rata-Rata"
                elif iq_value < 110:
                    values["KLASIFIKASI"] = "Rata-Rata"
                elif iq_value < 120:
                    values["KLASIFIKASI"] = "Diatas Rata-Rata"
                else:
                    values["KLASIFIKASI"] = "Superior"

            # Add PAPIKOSTICK values (kolom 18-38)
            papiko_fields = ["N", "G", "A", "L", "P", "I", "T", "V", "S", "B", "O", "X", 
                            "C", "D", "R", "Z", "E", "K", "F", "W"]  # Hapus "C (Coding)" dari list
        
            for i, field in enumerate(self.papikostick_inputs):
                if i < len(papiko_fields):
                    field_name = papiko_fields[i]
                    value = field.text().strip()
                    try:
                        if value:
                            float_value = float(value)
                            values[field_name] = float_value
                        else:
                            values[field_name] = "0"
                    except ValueError:
                        values[field_name] = value

                # Hitung C (Coding) berdasarkan nilai C
                if "C" in values:
                    try:
                        c_val = float(values["C"])
                        # Formula: =@IFS(AE4=1,9,AE4=2,8,AE4=3,7,AE4=4,6,AE4=5,5,AE4=6,4,AE4=7,3,AE4=8,2,AE4=9,1)
                        c_coding_map = {
                            1: 9,
                            2: 8,
                            3: 7,
                            4: 6,
                            5: 5,
                            6: 4,
                            7: 3,
                            8: 2,
                            9: 1
                        }
                        values["C (Coding)"] = c_coding_map.get(int(c_val), "")
                        print(f"DEBUG - C (Coding) dihitung otomatis: {values['C (Coding)']} dari nilai C: {c_val}")
                    except (ValueError, TypeError) as e:
                        print(f"Error menghitung C (Coding): {e}")
                        values["C (Coding)"] = ""

                # Calculate derived PAPIKOSTICK values (kolom 39-44)
                if all(values.get(field, "") != "" for field in ["S", "T"]):
                    values["NG"] = (float(values.get("S", 0)) + float(values.get("T", 0))) / 2

                if all(values.get(field, "") != "" for field in ["C", "D", "R"]):
                    values["CDR"] = (float(values.get("C", 0)) + float(values.get("D", 0)) + float(values.get("R", 0))) / 3

                if all(values.get(field, "") != "" for field in ["T", "V"]):
                    values["TV"] = (float(values.get("T", 0)) + float(values.get("V", 0))) / 2

                if all(values.get(field, "") != "" for field in ["P", "I"]):
                    values["PI"] = (float(values.get("P", 0)) + float(values.get("I", 0))) / 2

                if all(values.get(field, "") != "" for field in ["B", "S"]):
                    values["BS"] = (float(values.get("B", 0)) + float(values.get("S", 0))) / 2

                if all(values.get(field, "") != "" for field in ["Z", "K"]):
                    values["ZK"] = (float(values.get("Z", 0)) + float(values.get("K", 0))) / 2

                # Setelah perhitungan PAPIKOSTICK values dan sebelum menentukan mode add/edit, tambahkan:

                # Calculate Psikogram values (kolom 45-57)
                try:
                    # Logika Berpikir 1 (kolom 45) dan Logika Berpikir 1.1
                    if "SE / Konkrit Praktis" in values:
                        se_val = float(values.get("SE / Konkrit Praktis", 0))
                        # Untuk kolom asli
                        if se_val < 80:
                            values["Logika Berpikir 1"] = "R"
                        elif se_val < 100:
                            values["Logika Berpikir 1"] = "K"
                        elif se_val < 120:
                            values["Logika Berpikir 1"] = "C"
                        elif se_val < 140:
                            values["Logika Berpikir 1"] = "B"
                        else:
                            values["Logika Berpikir 1"] = "T"
                        
                        # Untuk kolom .1 (menggunakan referensi ke Sheet3)
                        values["Logika Berpikir 1.1"] = self.get_sheet3_reference("Logika Berpikir 1", values["Logika Berpikir 1"])

                    # Daya Analisa 3 (kolom 46) dan Daya Analisa 3.1
                    if "WA/ Verbal" in values:
                        wa_val = float(values.get("WA/ Verbal", 0))
                        # Untuk kolom asli
                        if wa_val < 80:
                            values["Daya Analisa 3"] = "R"
                        elif wa_val < 100:
                            values["Daya Analisa 3"] = "K"
                        elif wa_val < 120:
                            values["Daya Analisa 3"] = "C"
                        elif wa_val < 140:
                            values["Daya Analisa 3"] = "B"
                        else:
                            values["Daya Analisa 3"] = "T"
                        
                        # Untuk kolom .1
                        values["Daya Analisa 3.1"] = self.get_sheet3_reference("Daya Analisa 3", values["Daya Analisa 3"])

                    # Kemampuan Verbal 2 dam 4 (kolom 47) dan Kemampuan Verbal 2 dam 4.1
                    if "GE / Daya Abstraksi Verbal" in values:
                        ge_val = float(values.get("GE / Daya Abstraksi Verbal", 0))
                        # Untuk kolom asli
                        if ge_val < 80:
                            values["Kemampuan Verbal 2 dam 4"] = "R"
                        elif ge_val < 100:
                            values["Kemampuan Verbal 2 dam 4"] = "K"
                        elif ge_val < 120:
                            values["Kemampuan Verbal 2 dam 4"] = "C"
                        elif ge_val < 140:
                            values["Kemampuan Verbal 2 dam 4"] = "B"
                        else:
                            values["Kemampuan Verbal 2 dam 4"] = "T"
                        
                        # Untuk kolom .1
                        values["Kemampuan Verbal 2 dam 4.1"] = self.get_sheet3_reference("Kemampuan Verbal 2 dam 4", values["Kemampuan Verbal 2 dam 4"])

                    # Kemampuan Numerik 5 (kolom 48) dan Kemampuan Numerik 5.1
                    if " AN / Flexibilitas Pikir" in values:
                        an_val = float(values.get(" AN / Flexibilitas Pikir", 0))
                        # Untuk kolom asli
                        if an_val < 80:
                            values["Kemampuan Numerik 5"] = "R"
                        elif an_val < 100:
                            values["Kemampuan Numerik 5"] = "K"
                        elif an_val < 120:
                            values["Kemampuan Numerik 5"] = "C"
                        elif an_val < 140:
                            values["Kemampuan Numerik 5"] = "B"
                        else:
                            values["Kemampuan Numerik 5"] = "T"
                        
                        # Untuk kolom .1
                        values["Kemampuan Numerik 5.1"] = self.get_sheet3_reference("Kemampuan Numerik 5", values["Kemampuan Numerik 5"])

                    # Sistematika Kerja/ C D R (kolom 49) dan Sistematika Kerja/ C D R.1
                    if "CDR" in values:
                        cdr_val = float(values.get("CDR", 0))
                        # Untuk kolom asli
                        if cdr_val < 2:
                            values["Sistematika Kerja/ C D R"] = "R"
                        elif cdr_val < 4:
                            values["Sistematika Kerja/ C D R"] = "K"
                        elif cdr_val < 6:
                            values["Sistematika Kerja/ C D R"] = "C"
                        elif cdr_val < 9:
                            values["Sistematika Kerja/ C D R"] = "B"
                        elif cdr_val == 9:
                            values["Sistematika Kerja/ C D R"] = "T"
                        
                        # Untuk kolom .1
                        values["Sistematika Kerja/ C D R.1"] = self.get_sheet3_reference("Sistematika Kerja/ C D R", values["Sistematika Kerja/ C D R"])

                    # Orientasi Hasil/ N G (kolom 50) dan Orientasi Hasil/ N G.1
                    if all(values.get(field, "") != "" for field in ["N", "G"]):
                        n_val = float(values.get("N", 0))
                        g_val = float(values.get("G", 0))
                        ng_val = (n_val + g_val) / 2
                        # Untuk kolom asli
                        if ng_val < 2:
                            values["Orientasi Hasil/ N G"] = "R"
                        elif ng_val < 4:
                            values["Orientasi Hasil/ N G"] = "K"
                        elif ng_val < 6:
                            values["Orientasi Hasil/ N G"] = "C"
                        elif ng_val < 9:
                            values["Orientasi Hasil/ N G"] = "B"
                        elif ng_val == 9:
                            values["Orientasi Hasil/ N G"] = "T"
                        
                        # Untuk kolom .1
                        values["Orientasi Hasil/ N G.1"] = self.get_sheet3_reference("Orientasi Hasil/ N G", values["Orientasi Hasil/ N G"])

                    # Fleksibilitas/ T V (kolom 51) dan Fleksibilitas/ T V.1
                    if "TV" in values:
                        tv_val = float(values.get("TV", 0))
                        # Untuk kolom asli
                        if tv_val < 2:
                            values["Fleksibilitas/ T V"] = "R"
                        elif tv_val < 4:
                            values["Fleksibilitas/ T V"] = "K"
                        elif tv_val < 6:
                            values["Fleksibilitas/ T V"] = "C"
                        elif tv_val < 9:
                            values["Fleksibilitas/ T V"] = "B"
                        elif tv_val == 9:
                            values["Fleksibilitas/ T V"] = "T"
                        
                        # Untuk kolom .1
                        values["Fleksibilitas/ T V.1"] = self.get_sheet3_reference("Fleksibilitas/ T V", values["Fleksibilitas/ T V"])

                    # Motivasi Berprestasi/ A (kolom 52) dan Motivasi Berprestasi/ A.1
                    if "A" in values:
                        a_val = float(values.get("A", 0))
                        # Untuk kolom asli
                        if a_val < 2:
                            values["Motivasi Berprestasi/ A"] = "R"
                        elif a_val < 4:
                            values["Motivasi Berprestasi/ A"] = "K"
                        elif a_val < 6:
                            values["Motivasi Berprestasi/ A"] = "C"
                        elif a_val < 9:
                            values["Motivasi Berprestasi/ A"] = "B"
                        elif a_val == 9:
                            values["Motivasi Berprestasi/ A"] = "T"
                        
                        # Untuk kolom .1
                        values["Motivasi Berprestasi/ A.1"] = self.get_sheet3_reference("Motivasi Berprestasi/ A", values["Motivasi Berprestasi/ A"])

                    # Kerjasama/ P I (kolom 53) dan Kerjasama/ P I.1
                    if "PI" in values:
                        pi_val = float(values.get("PI", 0))
                        # Untuk kolom asli
                        if pi_val < 2:
                            values["Kerjasama/ P I"] = "R"
                        elif pi_val < 4:
                            values["Kerjasama/ P I"] = "K"
                        elif pi_val < 6:
                            values["Kerjasama/ P I"] = "C"
                        elif pi_val < 9:
                            values["Kerjasama/ P I"] = "B"
                        elif pi_val == 9:
                            values["Kerjasama/ P I"] = "T"
                        
                        # Untuk kolom .1
                        values["Kerjasama/ P I.1"] = self.get_sheet3_reference("Kerjasama/ P I", values["Kerjasama/ P I"])

                    # Keterampilan Interpersonal/ B S (kolom 54) dan Keterampilan Interpersonal/ B S.1
                    if "BS" in values:
                        bs_val = float(values.get("BS", 0))
                        # Untuk kolom asli
                        if bs_val < 2:
                            values["Keterampilan Interpersonal/ B S"] = "R"
                        elif bs_val < 4:
                            values["Keterampilan Interpersonal/ B S"] = "K"
                        elif bs_val < 6:
                            values["Keterampilan Interpersonal/ B S"] = "C"
                        elif bs_val < 9:
                            values["Keterampilan Interpersonal/ B S"] = "B"
                        elif bs_val == 9:
                            values["Keterampilan Interpersonal/ B S"] = "T"
                        
                        # Untuk kolom .1
                        values["Keterampilan Interpersonal/ B S.1"] = self.get_sheet3_reference("Keterampilan Interpersonal/ B S", values["Keterampilan Interpersonal/ B S"])

                    # Stabilitas Emosi/ E PHQ (kolom 55) dan Stabilitas Emosi/ E PHQ.1
                    if "E" in values:
                        e_val = float(values.get("E", 0))
                        # Untuk kolom asli
                        if e_val < 2:
                            values["Stabilitas Emosi/ E PHQ"] = "R"
                        elif e_val < 4:
                            values["Stabilitas Emosi/ E PHQ"] = "K"
                        elif e_val < 6:
                            values["Stabilitas Emosi/ E PHQ"] = "C"
                        elif e_val < 9:
                            values["Stabilitas Emosi/ E PHQ"] = "B"
                        elif e_val == 9:
                            values["Stabilitas Emosi/ E PHQ"] = "T"
                        
                        # Untuk kolom .1
                        values["Stabilitas Emosi/ E PHQ.1"] = self.get_sheet3_reference("Stabilitas Emosi/ E PHQ", values["Stabilitas Emosi/ E PHQ"])

                    # Pegembangan Diri/ W (kolom 56) dan Pegembangan Diri/ W.1
                    if "W" in values:
                        w_val = float(values.get("W", 0))
                        # Untuk kolom asli
                        if w_val < 2:
                            values["Pegembangan Diri/ W"] = "R"
                        elif w_val < 4:
                            values["Pegembangan Diri/ W"] = "K"
                        elif w_val < 6:
                            values["Pegembangan Diri/ W"] = "C"
                        elif w_val < 9:
                            values["Pegembangan Diri/ W"] = "B"
                        elif w_val == 9:
                            values["Pegembangan Diri/ W"] = "T"
                        
                        # Untuk kolom .1
                        values["Pegembangan Diri/ W.1"] = self.get_sheet3_reference("Pegembangan Diri/ W", values["Pegembangan Diri/ W"])

                    # Mengelola Perubahan/ Z K (kolom 57) dan Mengelola Perubahan/ Z K.1
                    if "ZK" in values:
                        zk_val = float(values.get("ZK", 0))
                        # Untuk kolom asli
                        if zk_val < 2:
                            values["Mengelola Perubahan/ Z K"] = "R"
                        elif zk_val < 4:
                            values["Mengelola Perubahan/ Z K"] = "K"
                        elif zk_val < 6:
                            values["Mengelola Perubahan/ Z K"] = "C"
                        elif zk_val < 9:
                            values["Mengelola Perubahan/ Z K"] = "B"
                        elif zk_val == 9:
                            values["Mengelola Perubahan/ Z K"] = "T"
                        
                        # Untuk kolom .1
                        values["Mengelola Perubahan/ Z K.1"] = self.get_sheet3_reference("Mengelola Perubahan/ Z K", values["Mengelola Perubahan/ Z K"])

                except Exception as e:
                    print(f"Error calculating psikogram values: {e}")
            
            # Determine action based on mode
            if mode == "add":
                row = self.table.rowCount()
                self.table.insertRow(row)
            elif mode == "edit":
                row = self.table.currentRow()
                if row < 0:
                    print("No row selected for editing")
                    QMessageBox.warning(self, "Warning", "Tidak ada baris yang dipilih untuk diedit")
                    return

            # Populate all columns
            for col_name in self.columns:
                col_idx = self.get_column_index(col_name)
                if col_idx >= 0:
                    value = values.get(col_name, "")
                    self.table.setItem(row, col_idx, QTableWidgetItem(str(value)))

            # Clear input fields after adding/updating
            for field in self.personal_inputs + self.ist_inputs + self.papikostick_inputs:
                if isinstance(field, QLineEdit):
                    field.clear()
                elif isinstance(field, QPushButton):
                    if field.text() not in ["Pilih Jenis Kelamin", "Pilih Tanggal"]:
                        field.setText("Pilih Tanggal" if "Tanggal" in field.text() else field.text())

            # Recalculate values for the row
            self.recalculate_values(row)

        except Exception as e:
            print(f"Error in add_or_update_row: {e}")
            import traceback
            traceback.print_exc()
            # QMessageBox.critical(self, "Error", f"Terjadi kesalahan: {str(e)}")

    def recalculate_values(self, row):
        try:
            # Simpan nilai W sebelum perhitungan
            original_w_value = self.get_w_value(row)
            print(f"DEBUG - Nilai W SEBELUM perhitungan: '{original_w_value}'")
            
            # Hitung nilai IQ = SUM(L4:P4)/5
            l_idx = self.get_column_index("SE / Konkrit Praktis")
            wa_idx = self.get_column_index("WA/ Verbal")
            an_idx = self.get_column_index(" AN / Flexibilitas Pikir")
            ge_idx = self.get_column_index("GE / Daya Abstraksi Verbal")
            ra_idx = self.get_column_index("RA / Berpikir Praktis")
            iq_idx = self.get_column_index("IQ ")
            
            if l_idx >= 0 and wa_idx >= 0 and an_idx >= 0 and ge_idx >= 0 and ra_idx >= 0 and iq_idx >= 0:
                l_val = self.convert_to_float(self.get_cell_text(row, l_idx))
                wa_val = self.convert_to_float(self.get_cell_text(row, wa_idx))
                an_val = self.convert_to_float(self.get_cell_text(row, an_idx))
                ge_val = self.convert_to_float(self.get_cell_text(row, ge_idx))
                ra_val = self.convert_to_float(self.get_cell_text(row, ra_idx))
                
                if l_val is not None and wa_val is not None and an_val is not None and ge_val is not None and ra_val is not None:
                    # Formula: =SUM(L4:P4)/5
                    iq_val = (l_val + wa_val + an_val + ge_val + ra_val) / 5
                    iq_item = QTableWidgetItem()
                    if iq_val.is_integer():
                        iq_item.setData(Qt.DisplayRole, int(iq_val))
                    else:
                        iq_item.setData(Qt.DisplayRole, iq_val)
                    self.table.setItem(row, iq_idx, iq_item)
                    print(f"Menghitung IQ: {iq_val}")
                    
                    # Hitung Unnamed: 16 = (M4+O4)/2
                    unnamed_16_idx = self.get_column_index("Unnamed: 16")
                    if unnamed_16_idx >= 0:
                        # Formula: =(M4+O4)/2 --> (AN + GE)/2
                        unnamed_16_val = (wa_val + ge_val) / 2
                        unnamed_16_item = QTableWidgetItem()
                        if unnamed_16_val.is_integer():
                            unnamed_16_item.setData(Qt.DisplayRole, int(unnamed_16_val))
                        else:
                            unnamed_16_item.setData(Qt.DisplayRole, unnamed_16_val)
                        self.table.setItem(row, unnamed_16_idx, unnamed_16_item)
                        print(f"Menghitung Unnamed: 16 = (AN + GE)/2: {unnamed_16_val}")
                    
                    # Hitung KLASIFIKASI berdasarkan IQ
                    if iq_val < 79:
                        iq_klasifikasi = "Rendah"
                    elif 79 <= iq_val < 90:
                        iq_klasifikasi = "Dibawah Rata-Rata"
                    elif 90 <= iq_val < 110:
                        iq_klasifikasi = "Rata-Rata"
                    elif 110 <= iq_val < 120:
                        iq_klasifikasi = "Diatas Rata-Rata"
                    else:
                        iq_klasifikasi = "Superior"
                        
                    # Cari indeks kolom KLASIFIKASI
                    klasifikasi_idx = self.get_column_index("KLASIFIKASI")
                    if klasifikasi_idx >= 0:
                        self.table.setItem(row, klasifikasi_idx, QTableWidgetItem(iq_klasifikasi))
                    print(f"Menghitung KLASIFIKASI: {iq_klasifikasi}")
            
            # Dapatkan nilai C untuk perhitungan C (Coding)
            c_idx = self.get_column_index("C")
            coding_idx = self.get_column_index("C (Coding)")
            
            if c_idx >= 0 and coding_idx >= 0:
                c_val = self.convert_to_float(self.get_cell_text(row, c_idx))
                
                # Formula: =IFS(AE4=1,9,AE4=2,8,AE4=3,7,AE4=4,6,AE4=5,5,AE4=6,4,AE4=7,3,AE4=8,2,AE4=9,1)
                if c_val is not None and 1 <= c_val <= 9:
                    c_coding = 10 - c_val
                    # Jika c_coding adalah integer, gunakan QTableWidgetItem dengan setData
                    c_coding_item = QTableWidgetItem()
                    if isinstance(c_coding, int) or (isinstance(c_coding, float) and c_coding.is_integer()):
                        c_coding_item.setData(Qt.DisplayRole, int(c_coding))
                    else:
                        c_coding_item.setData(Qt.DisplayRole, c_coding)
                    self.table.setItem(row, coding_idx, c_coding_item)
                    print(f"Formula IFS untuk kolom {coding_idx}: C (Coding) = 10 - C")
                    
            # Dapatkan nilai N dan G untuk perhitungan NG
            n_idx = self.get_column_index("N")
            g_idx = self.get_column_index("G")
            ng_idx = self.get_column_index("NG")
            
            if n_idx >= 0 and g_idx >= 0 and ng_idx >= 0:
                n_val = self.convert_to_float(self.get_cell_text(row, n_idx))
                g_val = self.convert_to_float(self.get_cell_text(row, g_idx))
                
                # Formula: =(S4+T4)/2
                if n_val is not None and g_val is not None:
                    ng_val = (n_val + g_val) / 2
                    ng_item = QTableWidgetItem()
                    if ng_val.is_integer():
                        ng_item.setData(Qt.DisplayRole, int(ng_val))
                    else:
                        ng_item.setData(Qt.DisplayRole, ng_val)
                    self.table.setItem(row, ng_idx, ng_item)
                    print(f"Formula normal untuk kolom {ng_idx}: NG = (N + G) / 2")
                    
            # Dapatkan nilai C, D, R untuk perhitungan CDR
            c_idx = self.get_column_index("C")
            d_idx = self.get_column_index("D")
            r_idx = self.get_column_index("R")
            cdr_idx = self.get_column_index("CDR")
            
            if c_idx >= 0 and d_idx >= 0 and r_idx >= 0 and cdr_idx >= 0:
                c_val = self.convert_to_float(self.get_cell_text(row, c_idx))
                d_val = self.convert_to_float(self.get_cell_text(row, d_idx))
                r_val = self.convert_to_float(self.get_cell_text(row, r_idx))
                
                # Formula: =(AE4+AG4+AH4)/3
                if c_val is not None and d_val is not None and r_val is not None:
                    cdr_val = (c_val + d_val + r_val) / 3
                    cdr_item = QTableWidgetItem()
                    if cdr_val.is_integer():
                        cdr_item.setData(Qt.DisplayRole, int(cdr_val))
                    else:
                        cdr_item.setData(Qt.DisplayRole, cdr_val)
                    self.table.setItem(row, cdr_idx, cdr_item)
                    print(f"Formula normal untuk kolom {cdr_idx}: CDR = (C + D + R) / 3")
                    
            # Dapatkan nilai T dan V untuk perhitungan TV
            t_idx = self.get_column_index("T")
            v_idx = self.get_column_index("V")
            tv_idx = self.get_column_index("TV")
            
            if t_idx >= 0 and v_idx >= 0 and tv_idx >= 0:
                t_val = self.convert_to_float(self.get_cell_text(row, t_idx))
                v_val = self.convert_to_float(self.get_cell_text(row, v_idx))
                
                # Formula: =(Y4+Z4)/2
                if t_val is not None and v_val is not None:
                    tv_val = (t_val + v_val) / 2
                    tv_item = QTableWidgetItem()
                    if tv_val.is_integer():
                        tv_item.setData(Qt.DisplayRole, int(tv_val))
                    else:
                        tv_item.setData(Qt.DisplayRole, tv_val)
                    self.table.setItem(row, tv_idx, tv_item)
                    print(f"Formula normal untuk kolom {tv_idx}: TV = (T + V) / 2")
                    
            # Dapatkan nilai P dan I untuk perhitungan PI
            p_idx = self.get_column_index("P")
            i_idx = self.get_column_index("I")
            pi_idx = self.get_column_index("PI")
            
            if p_idx >= 0 and i_idx >= 0 and pi_idx >= 0:
                p_val = self.convert_to_float(self.get_cell_text(row, p_idx))
                i_val = self.convert_to_float(self.get_cell_text(row, i_idx))
                
                # Formula: =(W4+X4)/2
                if p_val is not None and i_val is not None:
                    pi_val = (p_val + i_val) / 2
                    pi_item = QTableWidgetItem()
                    if pi_val.is_integer():
                        pi_item.setData(Qt.DisplayRole, int(pi_val))
                    else:
                        pi_item.setData(Qt.DisplayRole, pi_val)
                    self.table.setItem(row, pi_idx, pi_item)
                    print(f"Formula normal untuk kolom {pi_idx}: PI = (P + I) / 2")
                    
            # Dapatkan nilai B dan S untuk perhitungan BS
            b_idx = self.get_column_index("B")
            s_idx = self.get_column_index("S")
            bs_idx = self.get_column_index("BS")
            
            if b_idx >= 0 and s_idx >= 0 and bs_idx >= 0:
                b_val = self.convert_to_float(self.get_cell_text(row, b_idx))
                s_val = self.convert_to_float(self.get_cell_text(row, s_idx))
                
                # Formula: =(AA4+AB4)/2
                if b_val is not None and s_val is not None:
                    bs_val = (b_val + s_val) / 2
                    bs_item = QTableWidgetItem()
                    if bs_val.is_integer():
                        bs_item.setData(Qt.DisplayRole, int(bs_val))
                    else:
                        bs_item.setData(Qt.DisplayRole, bs_val)
                    self.table.setItem(row, bs_idx, bs_item)
                    print(f"Formula normal untuk kolom {bs_idx}: BS = (B + S) / 2")
                    
            # Dapatkan nilai Z dan K untuk perhitungan ZK
            z_idx = self.get_column_index("Z")
            k_idx = self.get_column_index("K")
            zk_idx = self.get_column_index("ZK")
            
            if z_idx >= 0 and k_idx >= 0 and zk_idx >= 0:
                z_val = self.convert_to_float(self.get_cell_text(row, z_idx))
                k_val = self.convert_to_float(self.get_cell_text(row, k_idx))
                
                # Formula: =(AI4+AK4)/2
                if z_val is not None and k_val is not None:
                    zk_val = (z_val + k_val) / 2
                    zk_item = QTableWidgetItem()
                    if zk_val.is_integer():
                        zk_item.setData(Qt.DisplayRole, int(zk_val))
                    else:
                        zk_item.setData(Qt.DisplayRole, zk_val)
                    self.table.setItem(row, zk_idx, zk_item)
                    print(f"Formula normal untuk kolom {zk_idx}: ZK = (Z + K) / 2")
            
            # Kembalikan nilai W ke nilai asli
            self.set_w_value(row, original_w_value)
            print(f"DEBUG - Nilai W SETELAH perhitungan: '{original_w_value}'")
                
        except Exception as e:
            print(f"Error dalam recalculate_values: {e}")
            import traceback
            traceback.print_exc()

    def get_papiko_index(self, code):
        """Helper function untuk mendapatkan indeks PAPIKOSTICK berdasarkan kode"""
        papiko_map = {
            'N': 0, 'G': 1, 'A': 2, 'L': 3, 'P': 4, 'I': 5, 'T': 6, 'V': 7,
            'S': 8, 'B': 9, 'O': 10, 'X': 11, 'C': 12, 'D': 13, 'R': 14,
            'Z': 15, 'E': 16, 'K': 17, 'F': 18, 'W': 38  # Ubah indeks W ke 38
        }
        return papiko_map.get(code)

    def get_cell_value(self, row, col):
        item = self.table.item(row, col)
        if item and item.text().strip():
            text = item.text()
            # Jika nilai adalah 'nan', kembalikan None
            if text.lower() == 'nan':
                return None
            try:
                return float(text)
            except ValueError:
                return None
        return None
        
    def get_cell_text(self, row, col):
        item = self.table.item(row, col)
        if item:
            # Jika nilai adalah 'nan', kembalikan string kosong
            if item.text().lower() == 'nan':
                return ""
            return item.text()
        return ""
        
    def convert_to_float(self, text):
        if text and text.strip():
            try:
                value = float(text)
                # Jika nilai bulat, konversi ke integer
                if value.is_integer():
                    return int(value)
                return value
            except ValueError:
                return None
        return None

    def delete_selected_row(self):
        selected_row = self.table.currentRow()
        if selected_row >= 0:
            self.table.removeRow(selected_row)
            self.table.resizeColumnsToContents()
            
    def print_selected_row(self):
        selected_row = self.table.currentRow()
        if selected_row >= 0:
            row_data = {}
            for col, column_name in enumerate(self.columns):
                item = self.table.item(selected_row, col)
                row_data[column_name] = item.text() if item else ""
            
            df_print = pd.DataFrame([row_data])
            temp_file = "temp_print.xlsx"
            df_print.to_excel(temp_file, index=False, engine="openpyxl")
            
            import os
            os.startfile(temp_file, "print")

    def save_to_excel(self):
        if not self.excel_file_path:
            return
        
        try:
            # Import yang diperlukan
            import os
            import shutil
            from datetime import datetime
            import openpyxl
            from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
            from openpyxl.utils.cell import get_column_letter
            
            # Buka workbook
            wb = openpyxl.load_workbook(self.excel_file_path, data_only=False)
            sheet = wb["Sheet1"] if "Sheet1" in wb.sheetnames else wb.active
            
            # Buat backup
            dir_path = os.path.dirname(self.excel_file_path)
            backup_dir = os.path.join(dir_path, "backup")
            os.makedirs(backup_dir, exist_ok=True)
            
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            backup_file = os.path.join(backup_dir, f"backup_{timestamp}_{os.path.basename(self.excel_file_path)}")
            shutil.copy2(self.excel_file_path, backup_file)
            
            # Cari baris header
            header_row = None
            for row in range(1, min(sheet.max_row + 1, 10)):
                for col in range(1, min(sheet.max_column + 1, 10)):
                    if sheet.cell(row=row, column=col).value == "No":
                        header_row = row
                        break
                if header_row:
                    break
            
            if not header_row:
                header_row = 3
            
            print("DEBUG - Scanning header row for columns...")
            
            # Buat mapping kolom
            col_mapping = {}
            unnamed_16_col = 17  # Kolom 17 untuk Unnamed: 16
            c_coding_col = None    # Untuk menyimpan kolom C (Coding)
            ae_col = None          # Untuk menyimpan kolom C (untuk rumus C Coding)
            wa_col = None          # Untuk rumus Unnamed: 16 (WA/Verbal)
            ge_col = None          # Untuk rumus Unnamed: 16 (GE)
            
            # Cetak semua header untuk debugging
            print("\nDEBUG - All headers in Excel:")
            for col in range(1, sheet.max_column + 1):
                header_value = sheet.cell(row=header_row, column=col).value
                if header_value:
                    print(f"Column {col}: '{header_value}'")
            
            print("\nDEBUG - All columns in UI table:")
            for col_name in self.columns:
                print(f"UI Column: '{col_name}'")
            
            # Scan header row
            for col in range(1, sheet.max_column + 1):
                header_value = sheet.cell(row=header_row, column=col).value
                if header_value:
                    header_str = str(header_value).strip()
                    col_mapping[header_str] = col
                    
                    # Deteksi kolom-kolom yang diperlukan
                    if header_str == "C (Coding)":
                        c_coding_col = col

                    elif header_str == "WA/ Verbal":
                        wa_col = col

                    elif header_str == "GE / Daya Abstraksi Verbal":
                        ge_col = col

                    elif header_str == "C":
                        ae_col = col

        
            
            # Simpan formula dari baris pertama data (jika ada)
            formulas = {}
            if sheet.max_row > header_row:
                first_data_row = header_row + 1
                for col in range(1, sheet.max_column + 1):
                    cell = sheet.cell(row=first_data_row, column=col)
                    if isinstance(cell.value, str) and cell.value.startswith('='):
                        formulas[col] = cell.value
            
            # Hapus data lama
            if sheet.max_row > header_row:
                sheet.delete_rows(header_row + 1, sheet.max_row - header_row)
            
            # Simpan style header yang penting saja
            header_styles = {}
            for col in range(1, sheet.max_column + 1):
                cell = sheet.cell(row=header_row, column=col)
                if cell.has_style:
                    # Hanya simpan fill dan border
                    border = Border(
                        left=cell.border.left,
                        right=cell.border.right,
                        top=cell.border.top,
                        bottom=cell.border.bottom
                    )
                    
                    fill = None
                    if cell.fill and cell.fill.fill_type:
                        fill = PatternFill(
                            fill_type=cell.fill.fill_type,
                            start_color=cell.fill.start_color,
                            end_color=cell.fill.end_color
                        )
                    
                    header_styles[col] = {
                        'border': border,
                        'fill': fill
                    }
            
            # Formula untuk kolom psikogram (46-58)
            psikogram_formulas = {
                46: "=IF(L{row}<80,\"R\",IF(L{row}<100,\"K\",IF(L{row}<120,\"C\",IF(L{row}<140,\"B\",\"T\"))))",  # Logika Berpikir 1
                47: "=IF(M{row}<80,\"R\",IF(M{row}<100,\"K\",IF(M{row}<120,\"C\",IF(M{row}<140,\"B\",\"T\"))))",  # Daya Analisa 3
                48: "=IF(O{row}<80,\"R\",IF(O{row}<100,\"K\",IF(O{row}<120,\"C\",IF(O{row}<140,\"B\",\"T\"))))",  # Kemampuan Verbal 2 dam 4
                49: "=IF(N{row}<80,\"R\",IF(N{row}<100,\"K\",IF(N{row}<120,\"C\",IF(N{row}<140,\"B\",\"T\"))))",  # Kemampuan Numerik 5
                50: "=IFS(AO{row}<2,\"R\",AO{row}<4,\"K\",AO{row}<6,\"C\",AO{row}<9,\"B\",AO{row}=9,\"T\")",  # Sistematika Kerja/ C D R
                51: "=IFS(AN{row}<2,\"R\",AN{row}<4,\"K\",AN{row}<6,\"C\",AN{row}<9,\"B\",AN{row}=9,\"T\")",  # Orientasi Hasil/ N G
                52: "=IFS(AP{row}<2,\"R\",AP{row}<4,\"K\",AP{row}<6,\"C\",AP{row}<9,\"B\",AP{row}=9,\"T\")",  # Fleksibilitas/ T V
                53: "=IFS(U{row}<2,\"R\",U{row}<4,\"K\",U{row}<6,\"C\",U{row}<9,\"B\",U{row}=9,\"T\")",  # Motivasi Berprestasi/ A
                54: "=IFS(AQ{row}<2,\"R\",AQ{row}<4,\"K\",AQ{row}<6,\"C\",AQ{row}<9,\"B\",AQ{row}=9,\"T\")",  # Kerjasama/ P I
                55: "=IFS(AR{row}<2,\"R\",AR{row}<4,\"K\",AR{row}<6,\"C\",AR{row}<9,\"B\",AR{row}=9,\"T\")",  # Keterampilan Interpersonal/ B S
                56: "=IFS(AJ{row}<2,\"R\",AJ{row}<4,\"K\",AJ{row}<6,\"C\",AJ{row}<9,\"B\",AJ{row}=9,\"T\")",  # Stabilitas Emosi/ E PHQ
                57: "=IFS(AM{row}<2,\"R\",AM{row}<4,\"K\",AM{row}<6,\"C\",AM{row}<9,\"B\",AM{row}=9,\"T\")",  # Pegembangan Diri/ W
                58: "=IFS(AS{row}<2,\"R\",AS{row}<4,\"K\",AS{row}<6,\"C\",AS{row}<9,\"B\",AS{row}=9,\"T\")",  # Mengelola Perubahan/ Z K
            }
            
            # Tambahkan mapping kolom psikogram
            psikogram_column_mapping = {
                "Logika Berpikir 1": 46,
                "Daya Analisa 3": 47,
                "Kemampuan Verbal 2 dam 4": 48,
                "Kemampuan Numerik 5": 49,
                "Sistematika Kerja/ C D R": 50,
                "Orientasi Hasil/ N G": 51,
                "Fleksibilitas/ T V": 52,
                "Motivasi Berprestasi/ A": 53,
                "Kerjasama/ P I": 54,
                "Keterampilan Interpersonal/ B S": 55,
                "Stabilitas Emosi/ E PHQ": 56,
                "Pegembangan Diri/ W": 57,
                "Mengelola Perubahan/ Z K": 58
            }
            
            # Tulis data baru
            for row_idx, row in enumerate(range(self.table.rowCount()), start=1):
                target_row = header_row + row_idx
                
                # Dictionary untuk menyimpan nilai psikogram untuk baris ini
                psikogram_values = {}
                
                # Simpan nilai WA dan GE untuk perhitungan Unnamed: 16
                wa_value = None
                ge_value = None
                
                # Pertama, kumpulkan nilai WA dan GE
                for col_idx, col_name in enumerate(self.columns):
                    item = self.table.item(row, col_idx)
                    value = item.text() if item else ""
                    
                    # Debug untuk memastikan kolom terisi
                    print(f"DEBUG - Menulis nilai '{value}' ke kolom '{col_name}' di baris {target_row}")
                    
                    # Cari kolom yang sesuai di Excel
                    excel_col = None
                    
                    # Bersihkan nama kolom dari spasi di awal dan akhir
                    col_name_clean = col_name.strip()
                    
                    # Penanganan khusus untuk kolom dengan suffix .1
                    if col_name_clean.endswith('.1'):
                        # Hapus suffix .1 untuk mencari di Excel
                        base_name = col_name_clean[:-2]  # Hapus '.1' dari nama
                        # Cari di col_mapping dengan nama dasar
                        for header, col in col_mapping.items():
                            if header.strip() == base_name:
                                excel_col = col
                                print(f"DEBUG - Matched UI column '{col_name}' to Excel column '{header}' at position {col}")
                                break
                    else:
                        # Penanganan normal untuk kolom lain
                        # Coba cari dengan nama yang sudah dibersihkan
                        for header, col in col_mapping.items():
                            if header.strip() == col_name_clean:
                                excel_col = col
                                print(f"DEBUG - Matched regular column '{col_name}' to Excel column '{header}' at position {col}")
                                break
                        
                        # Khusus untuk Unnamed: 16
                        if col_name == "Unnamed: 16":
                            excel_col = unnamed_16_col
                    
                    if excel_col is None:
                        print(f"DEBUG - No match found for UI column '{col_name}' (cleaned: '{col_name_clean}') in Excel")
                        continue
                    
                    # Tulis nilai ke Excel
                    target_cell = sheet.cell(row=target_row, column=excel_col)
                    
                    # Terapkan style dari header (hanya border dan fill)
                    if excel_col in header_styles:
                        style = header_styles[excel_col]
                        target_cell.border = style['border']
                        if style['fill']:
                            target_cell.fill = style['fill']
                    
                    # Penanganan khusus untuk kolom Unnamed: 16
                    if col_name == "Unnamed: 16" and wa_col and ge_col:
                        wa_col_letter = get_column_letter(wa_col)
                        ge_col_letter = get_column_letter(ge_col)
                        formula = f"=({wa_col_letter}{target_row}+{ge_col_letter}{target_row})/2"
                        target_cell.value = formula
                        print(f"DEBUG - Setting formula for Unnamed: 16: {formula}")
                    
                    # Penanganan khusus untuk kolom C (Coding)
                    elif excel_col == c_coding_col and ae_col:
                        # Rumus untuk C (Coding)
                        ae_col_letter = get_column_letter(ae_col)
                        target_cell.value = f"=IFS({ae_col_letter}{target_row}=1,9,{ae_col_letter}{target_row}=2,8,{ae_col_letter}{target_row}=3,7,{ae_col_letter}{target_row}=4,6,{ae_col_letter}{target_row}=5,5,{ae_col_letter}{target_row}=6,4,{ae_col_letter}{target_row}=7,3,{ae_col_letter}{target_row}=8,2,{ae_col_letter}{target_row}=9,1)"
                    
                    # Penanganan khusus untuk kolom psikogram
                    elif col_name_clean in psikogram_column_mapping:
                        excel_col = psikogram_column_mapping[col_name_clean]
                        if excel_col in psikogram_formulas:
                            formula = psikogram_formulas[excel_col].replace("{row}", str(target_row))
                            target_cell = sheet.cell(row=target_row, column=excel_col)
                            target_cell.value = formula
                            print(f"DEBUG - Setting formula for psikogram column {col_name_clean} at position {excel_col}: {formula}")
                            psikogram_values[excel_col] = value
                    
                    # Penanganan khusus untuk kolom referensi (.1)
                    elif col_name_clean.endswith('.1'):
                        base_name = col_name_clean[:-2]
                        if base_name in psikogram_column_mapping:
                            psikogram_col = psikogram_column_mapping[base_name]
                            excel_col = psikogram_col + 13  # Kolom referensi dimulai 13 kolom setelah psikogram
                            
                            # Dapatkan huruf kolom untuk referensi
                            source_col_letter = get_column_letter(psikogram_col)
                            
                            # Buat formula referensi ke Sheet3 berdasarkan nilai di kolom psikogram
                            formula = (
                                f'=IF({source_col_letter}{target_row}="R",Sheet3!$F$30,'
                                f'IF({source_col_letter}{target_row}="K",Sheet3!$F$31,'
                                f'IF({source_col_letter}{target_row}="C",Sheet3!$F$32,'
                                f'IF({source_col_letter}{target_row}="B",Sheet3!$F$33,'
                                f'IF({source_col_letter}{target_row}="T",Sheet3!$F$34)))))'
                            )
                            
                            # Sesuaikan referensi sel Sheet3 berdasarkan jenis kolom
                            sheet3_refs = {
                                "Logika Berpikir 1": ("$F$30", "$F$31", "$F$32", "$F$33", "$F$34"),
                                "Daya Analisa 3": ("$F$35", "$F$36", "$F$37", "$F$38", "$F$39"),
                                "Kemampuan Verbal 2 dam 4": ("$F$45", "$F$46", "$F$47", "$F$48", "$F$49"),
                                "Kemampuan Numerik 5": ("$F$40", "$F$41", "$F$42", "$F$43", "$F$44"),
                                "Sistematika Kerja/ C D R": ("$F$60", "$F$61", "$F$62", "$F$63", "$F$64"),
                                "Orientasi Hasil/ N G": ("$F$50", "$F$51", "$F$52", "$F$53", "$F$54"),
                                "Fleksibilitas/ T V": ("$F$55", "$F$56", "$F$57", "$F$58", "$F$59"),
                                "Motivasi Berprestasi/ A": ("$F$65", "$F$66", "$F$67", "$F$68", "$F$69"),
                                "Kerjasama/ P I": ("$F$70", "$F$71", "$F$72", "$F$73", "$F$74"),
                                "Keterampilan Interpersonal/ B S": ("$F$75", "$F$76", "$F$77", "$F$78", "$F$79"),
                                "Stabilitas Emosi/ E PHQ": ("$F$80", "$F$81", "$F$82", "$F$83", "$F$84"),
                                "Pegembangan Diri/ W": ("$F$85", "$F$86", "$F$87", "$F$88", "$F$89"),
                                "Mengelola Perubahan/ Z K": ("$F$90", "$F$91", "$F$92", "$F$93", "$F$94")
                            }
                            
                            if base_name in sheet3_refs:
                                refs = sheet3_refs[base_name]
                                formula = (
                                    f'=IF({source_col_letter}{target_row}="R",Sheet3!{refs[0]},'
                                    f'IF({source_col_letter}{target_row}="K",Sheet3!{refs[1]},'
                                    f'IF({source_col_letter}{target_row}="C",Sheet3!{refs[2]},'
                                    f'IF({source_col_letter}{target_row}="B",Sheet3!{refs[3]},'
                                    f'IF({source_col_letter}{target_row}="T",Sheet3!{refs[4]})))))'
                                )
                            
                            target_cell = sheet.cell(row=target_row, column=excel_col)
                            target_cell.value = formula
                            print(f"DEBUG - Setting formula for reference column {col_name_clean} at position {excel_col}: {formula}")
                    
                    # Untuk kolom lain, gunakan formula jika ada atau nilai biasa
                    elif excel_col in formulas:
                        formula = formulas[excel_col]
                        formula = formula.replace(str(header_row + 1), str(target_row))
                        target_cell.value = formula
                    else:
                        try:
                            if value.replace('.', '', 1).replace('-', '', 1).isdigit():
                                num_val = float(value)
                                if num_val.is_integer():
                                    target_cell.value = int(num_val)
                                else:
                                    target_cell.value = num_val
                            else:
                                target_cell.value = value
                        except:
                            target_cell.value = value

                # Pastikan nilai benar-benar ditulis
                print(f"DEBUG - Nilai yang ditulis ke Excel: '{target_cell.value}' di sel {get_column_letter(excel_col)}{target_row}")

            # Sesuaikan tinggi baris
            for row in range(header_row + 1, sheet.max_row + 1):
                sheet.row_dimensions[row].height = 15

            try:
                # Simpan workbook
                wb.save(self.excel_file_path)
                wb.close()
                QMessageBox.information(self, "Sukses", f"Data berhasil disimpan ke {self.excel_file_path}")
                print(f"DEBUG - Workbook saved successfully to {self.excel_file_path}")
            except Exception as e:
                print(f"ERROR - Failed to save workbook: {e}")
                QMessageBox.critical(self, "Error", f"Gagal menyimpan file Excel: {e}")
                
                # Coba kembalikan dari backup jika ada error
                try:
                    latest_backup = max(
                        [f for f in os.listdir(backup_dir) if f.startswith("backup_") and f.endswith(os.path.basename(self.excel_file_path))],
                        key=lambda x: os.path.getmtime(os.path.join(backup_dir, x))
                    )
                    backup_path = os.path.join(backup_dir, latest_backup)
                    shutil.copy2(backup_path, self.excel_file_path)
                    QMessageBox.information(self, "Pemulihan", f"File dikembalikan dari backup: {latest_backup}")
                except Exception as restore_error:
                    print(f"Error restoring from backup: {restore_error}")
            
        except Exception as e:
            print(f"Error saving to Excel: {e}")
            import traceback
            traceback.print_exc()
            QMessageBox.critical(self, "Error", f"Gagal menyimpan file Excel: {e}")
            
            # Coba kembalikan dari backup jika ada error
            try:
                latest_backup = max(
                    [f for f in os.listdir(backup_dir) if f.startswith("backup_") and f.endswith(os.path.basename(self.excel_file_path))],
                    key=lambda x: os.path.getmtime(os.path.join(backup_dir, x))
                )
                backup_path = os.path.join(backup_dir, latest_backup)
                shutil.copy2(backup_path, self.excel_file_path)
                QMessageBox.information(self, "Pemulihan", f"File dikembalikan dari backup: {latest_backup}")
            except Exception as restore_error:
                print(f"Error restoring from backup: {restore_error}")

    def get_column_index(self, column_name):
        # Search for exact match first
        try:
            return self.columns.index(column_name)
        except ValueError:
            # If exact match not found, try case insensitive match
            for i, col in enumerate(self.columns):
                if col.lower() == column_name.lower():
                    return i
                # For special case like "IQ" and "IQ " (with space)
                if column_name == "IQ" and col == "IQ ":
                    return i
                if column_name == "IQ " and col == "IQ":
                    return i
            # If still not found, try to find by partial match
            for i, col in enumerate(self.columns):
                if column_name.lower() in col.lower():
                    return i
            # If nothing found
            return -1

    def preview_pdf(self):
        try:
            # Check if Excel file has been loaded
            if not hasattr(self, 'excel_file_path') or not self.excel_file_path:
                QMessageBox.warning(self, "Warning", "Please load an Excel file first!")
                return

            # Check if a row is selected
            selected_row = self.table.currentRow()
            if selected_row < 0:
                QMessageBox.warning(self, "Warning", "Please select a row to preview first!")
                return

            # Get absolute path for logo
            import os
            import base64
            from datetime import datetime
            current_dir = os.path.dirname(os.path.abspath(__file__))
            logo_path = os.path.join(current_dir, "logo.png")
            
            # Baca file logo dan encode ke base64
            try:
                # Coba cara pertama dengan path relatif
                if os.path.exists(logo_path):
                    with open(logo_path, "rb") as image_file:
                        logo_base64 = base64.b64encode(image_file.read()).decode('utf-8')
                        # Konversi ke data URL untuk digunakan dalam HTML dengan format PNG
                        logo_data_url = f"data:image/png;base64,{logo_base64}"
                        print(f"Logo encoded successfully. Path: {logo_path}")
                else:
                    # Coba dengan direktori saat ini
                    logo_path_alt = "logo.png"
                    if os.path.exists(logo_path_alt):
                        with open(logo_path_alt, "rb") as image_file:
                            logo_base64 = base64.b64encode(image_file.read()).decode('utf-8')
                            logo_data_url = f"data:image/png;base64,{logo_base64}"
                            print(f"Logo encoded successfully using alternative path: {logo_path_alt}")
                    else:
                        # Jika masih gagal, periksa direktori di atas
                        logo_path_parent = os.path.join(os.path.dirname(current_dir), "logo.png")
                        if os.path.exists(logo_path_parent):
                            with open(logo_path_parent, "rb") as image_file:
                                logo_base64 = base64.b64encode(image_file.read()).decode('utf-8')
                                logo_data_url = f"data:image/png;base64,{logo_base64}"
                                print(f"Logo encoded successfully using parent directory: {logo_path_parent}")
                        else:
                            # Fallback jika semua upaya gagal
                            print(f"Logo not found in any of these locations: {logo_path}, {logo_path_alt}, {logo_path_parent}")
                            logo_data_url = ""
            except Exception as e:
                print(f"Error encoding logo: {e}")
                # Fallback jika logo tidak bisa dibaca
                logo_data_url = ""
            
            # Get column indices and data
            iq_col = self.get_column_index("IQ ")
            nama_col = self.get_column_index("Nama Peserta")
            tgl_lahir_col = self.get_column_index("TGL Lahir")

            # Get data from selected row
            iq_val = self.table.item(selected_row, iq_col)
            nama_val = self.table.item(selected_row, nama_col)
            tgl_lahir_val = self.table.item(selected_row, tgl_lahir_col)

            # Initialize row_data from the selected row
            row_data = {}
            for col, column_name in enumerate(self.columns):
                item = self.table.item(selected_row, col)
                row_data[column_name] = item.text() if item else ""

            # Debug prints
            print(f"Selected Row: {selected_row}")
            for key, value in row_data.items():
                print(f"{key}: {value}")

            nama = nama_val.text() if nama_val else ""
            tgl_lahir = tgl_lahir_val.text() if tgl_lahir_val else ""

            # Ensure iq_val is converted to float correctly
            try:
                iq_value_numeric = float(iq_val.text()) if iq_val and iq_val.text().strip() else 0.0
            except ValueError:
                iq_value_numeric = 0.0

            # For debugging
            print(f"IQ Value from table: {iq_value_numeric}")

            # Create HTML content
            html_content = """
            <html>
            <head>
                <meta charset="UTF-8">
                <style>
                    @page {
                        size: A4;
                        margin: 1cm;
                    }
                    body { 
                        font-family: Arial, sans-serif;
                        padding: 20px;
                        width: 21cm;
                        min-height: 29.7cm;
                        margin: 0 auto;
                        background: white;
                    }
                    @media print {
                        body {
                            width: auto;
                            height: auto;
                            margin: 0;
                            padding: 0;
                        }
                        .page {
                            width: 21cm;
                            min-height: 29.7cm;
                            padding: 1cm;
                            margin: 0;
                            page-break-after: always;
                        }
                        .page:last-child {
                            page-break-after: avoid;
                        }
                        .page-break {
                            page-break-before: always;
                            margin: 0;
                            padding: 0;
                        }
                    }
                    .header { text-align: center; margin-bottom: 20px; }
                    .title { font-size: 16px; font-weight: bold; margin: 10px 0; }
                    .info-table { width: 100%; margin-bottom: 15px; border-spacing: 0; table-layout:fixed;  }
                    .info-table td { padding: 3px; vertical-align: top; width: 25%;  }
                    .main-table, .psikogram { 
                        width: 100%; 
                        border-collapse: collapse; 
                        margin-bottom: 20px;
                    }
                    .main-table th, .main-table td, .psikogram th, .psikogram td { 
                        border: 1px solid black; 
                        padding: 8px; 
                        text-align: center;
                    }
                    .psikogram td { text-align: left; }
                    .main-table th, .psikogram th { background-color: #f2f2f2; }
                    .center-text { text-align: center; }
                    .footer { text-align: center; font-style: italic; margin-top: 20px; }
                    .psikogram th, .psikogram td { 
                        padding: 10px; 
                        font-size: 12px; 
                        vertical-align: middle;
                    }
                </style>
            </head>
            """

            # Add personal info
            # Get test date from Excel
            tgl_test_col = self.get_column_index("Tgl Test")
            tgl_test_val = self.table.item(selected_row, tgl_test_col)
            tanggal_tes = tgl_test_val.text() if tgl_test_val else datetime.now().strftime("%d %B %Y")

            # Convert and format birth date from Excel
            try:
                # Try parse with DD/MM/YYYY format
                tgl_lahir_obj = datetime.strptime(tgl_lahir, "%d/%m/%Y")
            except ValueError:
                try:
                    # If failed, try parse with YYYY-MM-DD format
                    tgl_lahir_obj = datetime.strptime(tgl_lahir, "%Y-%m-%d")
                except ValueError:
                    # If still failed, use current date as fallback
                    print(f"Error: Date format '{tgl_lahir}' not recognized. Using current date.")
                    tgl_lahir_obj = datetime.now()
                    
            tgl_lahir_formatted = tgl_lahir_obj.strftime("%d %B %Y")

            # Convert and format test date from Excel
            try:
                # Try parse with DD/MM/YYYY format
                tgl_test_obj = datetime.strptime(tanggal_tes, "%d/%m/%Y")
            except ValueError:
                try:
                    # If failed, try parse with YYYY-MM-DD format
                    tgl_test_obj = datetime.strptime(tanggal_tes, "%Y-%m-%d")
                except ValueError:
                    # If still failed, use current date as fallback
                    print(f"Error: Date format '{tanggal_tes}' not recognized. Using current date.")
                    tgl_test_obj = datetime.now()
                    
            tanggal_tes_formatted = tgl_test_obj.strftime("%d %B %Y")

            # Get company name from Excel
            nama_pt_col = self.get_column_index("Nama PT")
            nama_pt_val = self.table.item(selected_row, nama_pt_col)
            nama_perusahaan = nama_pt_val.text() if nama_pt_val else "PT. BAM"

            html_content += f"""
            <div style="width: 100%; margin: 0 auto;">
                <div style="display: flex; justify-content: space-between; align-items: center; margin-bottom: 20px;">
                    <div style="width: 150px;">
                        <img src="{logo_data_url}" alt="Logo" style="width: 100%; height: auto;">
                    </div>
                    <div style="text-align: center; flex-grow: 1;">
                        <div style="font-size: 15px; font-weight: bold; color: #1f4e79;">HASIL PEMERIKSAAN PSIKOLOGIS</div>
                        <div style="font-size: 17px; font-weight: bold; color: #1f4e79;">(Asesmen Intelegensi, Kepribadian dan Minat)</div>
                    </div>
                    <div style="text-align: right; font-size: 12px;">
                        <div style="font-weight: bold; color: #1f4e79;">RAHASIA</div>
                        <div style="color: #1f4e79;">No. {row_data.get('No', '')}{row_data.get('No Tes', '')}</div>
                    </div>
                </div>

                <table class="info-table" style="margin-bottom: 20px; border-spacing: 0; font-size: 12px; width: 100%;">
                    <tr>
                        <td width="15%" style="padding: 4px 0; color: #c45911; font-weight: bold; text-align: left;">NAMA</td>
                        <td width="35%" style="padding: 4px 0; color: #c45911; font-weight: bold; text-align: left;">: {nama}</td>
                        <td width="15%" style="padding: 4px 0; color: #c45911; font-weight: bold; text-align: left;">PERUSAHAAN</td>
                        <td width="35%" style="padding: 4px 0; color: #c45911; font-weight: bold; text-align: left;">: {nama_perusahaan}</td>
                    </tr>
                    <tr>
                        <td style="padding: 4px 0; color: #c45911; font-weight: bold; text-align: left;">TANGGAL LAHIR</td>
                        <td style="padding: 4px 0; color: #c45911; font-weight: bold; text-align: left;">: {tgl_lahir_formatted}</td>
                        <td style="padding: 4px 0; color: #c45911; font-weight: bold; text-align: left;">TANGGAL TES</td>
                        <td style="padding: 4px 0; color: #c45911; font-weight: bold; text-align: left;">: {tanggal_tes_formatted}</td>
                    </tr>
                    <tr>
                        <td style="padding: 4px 0; color: #c45911; font-weight: bold; text-align: left;">PEMERIKSA</td>
                        <td style="padding: 4px 0; color: #c45911; font-weight: bold; text-align: left;">: Chitra Ananda Mulia, M.Psi., Psikolog</td>
                        <td style="padding: 4px 0; color: #c45911; font-weight: bold; text-align: left;">LEMBAGA</td>
                        <td style="padding: 4px 0; color: #c45911; font-weight: bold; text-align: left;">: BEHAVYOURS</td>
                    </tr>
                    <tr>
                        <td style="padding: 4px 0; color: #c45911; font-weight: bold; text-align: left;">ALAMAT LEMBAGA</td>
                        <td colspan="3" style="padding: 4px 0; color: #c45911; font-weight: bold; text-align: left;">: Jl. Patal Senayan No.01</td>
                    </tr>
                </table>
            </div>
            """
            # Pastikan iq_value adalah numerik
            try:
                iq_value_numeric = float(iq_val.text())
            except ValueError:
                iq_value_numeric = 0 
            # Add IQ Classification table
            html_content += f"""
                    <div style="width: 100%; margin: 0 auto;">
                        <table style="width: 100%; border-collapse: separate; border-spacing: 0 0;">
                            <tr>
                                <td style="width: 25%; padding-right: 15px; vertical-align: top;">
                                    <table style="width: 100%; border-collapse: collapse; border: 1px solid black;">
                                        <tr>
                                            <th style="border-bottom: 1px solid black; padding: 8px; text-align: center; background-color: #f7caac;">KECERDASAN UMUM</th>
                                        </tr>
                                        <tr>
                                            <td>
                                                <table style="width: 100%; border-collapse: collapse;">
                                                    <tr>
                                                        <td style="border-right: 1px solid black; padding: 8px; text-align: left;">Taraf<br>Kecerdasan<br>IQ</td>
                                                        <td style="padding: 8px; text-align: center;">{iq_val.text()}</td>
                                                    </tr>
                                                </table>
                                            </td>
                                        </tr>
                                    </table>
                                </td>
                                <td style="width: 75%; vertical-align: top;">
                                    <table style="width: 100%; border-collapse: collapse;">
                                        <tr>
                                            <th colspan="5" style="border: 1px solid black; padding: 4px; text-align: center; background-color: #ffceb4;">KLASIFIKASI KECERDASAN IQ</th>
                                        </tr>
                                        <tr>
                                            <td style="border: 1px solid black; padding: 4px; text-align: center;">Rendah</td>
                                            <td style="border: 1px solid black; padding: 4px; text-align: center;">Dibawah Rata-Rata</td>
                                            <td style="border: 1px solid black; padding: 4px; text-align: center;">Rata-Rata</td>
                                            <td style="border: 1px solid black; padding: 4px; text-align: center;">Diatas Rata-Rata</td>
                                            <td style="border: 1px solid black; padding: 4px; text-align: center;">Superior</td>
                                        </tr>
                                        <tr>
                                            <td style="border: 1px solid black; padding: 4px; text-align: center;">&lt; 79</td>
                                            <td style="border: 1px solid black; padding: 4px; text-align: center;">80 - 89</td>
                                            <td style="border: 1px solid black; padding: 4px; text-align: center;">90 - 109</td>
                                            <td style="border: 1px solid black; padding: 4px; text-align: center;">110 - 119</td>
                                            <td style="border: 1px solid black; padding: 4px; text-align: center;">&gt; 120</td>
                                        </tr>
                                    </table>
                                </td>
                            </tr>
                        </table>   
                    </div>     
            """

            # Add IQ data
            iq_val = self.table.item(selected_row, self.get_column_index("IQ "))
            iq_class = self.table.item(selected_row, self.get_column_index("KLASIFIKASI"))
            if iq_val and iq_class:

                # Define a function to determine the position of 'X'
                def get_x_position(value):
                    positions = {
                        "R": ["X", "", "", "", ""],
                        "K": ["", "X", "", "", ""],
                        "C": ["", "", "X", "", ""],
                        "B": ["", "", "", "X", ""],
                        "T": ["", "", "", "", "X"]
                    }
                    return positions.get(value, ["", "", "", "", ""])

                # Get values for each psychological aspect
                aspects = {
                    "logika_berpikir": row_data.get("Logika Berpikir 1", ""),  # Sesuai kolom asli
                    "daya_analisa": row_data.get("Daya Analisa 3", ""),  # Sesuai kolom asli
                    "kemampuan_verbal": row_data.get("Kemampuan Verbal 2 dam 4", ""),  # Sesuai kolom asli
                    "kemampuan_numerik": row_data.get("Kemampuan Numerik 5", ""),  # Sesuai kolom asli
                    "sistematika_kerja": row_data.get("Sistematika Kerja/ C D R", ""),  # Sesuai kolom asli
                    "orientasi_hasil": row_data.get("Orientasi Hasil/ N G", ""),  # Sesuai kolom asli
                    "fleksibilitas": row_data.get("Fleksibilitas/ T V", ""),  # Sesuai kolom asli
                    "motivasi_berprestasi": row_data.get("Motivasi Berprestasi/ A", ""),  # Sesuai kolom asli
                    "kerjasama": row_data.get("Kerjasama/ P I", ""),  # Sesuai kolom asli
                    "keterampilan_interpersonal": row_data.get("Keterampilan Interpersonal/ B S", ""),  # Sesuai kolom asli
                    "stabilitas_emosi": row_data.get("Stabilitas Emosi/ E PHQ", ""),  # Sesuai kolom asli
                    "pengembangan_diri": row_data.get("Pegembangan Diri/ W", ""),  # Sesuai kolom asli
                    "mengelola_perubahan": row_data.get("Mengelola Perubahan/ Z K", "")  # Sesuai kolom asli
                }

                # Define common cell styles
                header_style = 'text-align: center; padding: 8px; background-color: #deeaf6; border: 1px solid black;'
                cell_style = 'text-align: center; font-weight: bold;'
                section_header_style = 'background-color: #fbe4d5; text-align: center; font-weight: bold;'
                
                # Add psikogram table with dynamic 'X' positions
                html_content += f"""
                    <table class="psikogram" style="width: 100%; margin-top: 20px; border-collapse: collapse; border: 1px solid black;">
                        <tr>
                            <th colspan="8" style="{header_style}">PSIKOGRAM</th>
                        </tr>
                        <tr>
                            <th style="width: 5%; {header_style}">NO</th>
                            <th style="width: 15%; {header_style}">ASPEK<br>PSIKOLOGIS</th>
                            <th style="width: 40%; {header_style}">DEFINISI</th>
                            <th style="width: 8%; {header_style}">R</th>
                            <th style="width: 8%; {header_style}">K</th>
                            <th style="width: 8%; {header_style}">C</th>
                            <th style="width: 8%; {header_style}">B</th>
                            <th style="width: 8%; {header_style}">T</th>
                        </tr>

                        <tr><td colspan="8" style="{section_header_style}">KEMAMPUAN INTELEKTUAL</td></tr>
                        
                        <!-- Logika Berpikir -->
                        <tr>
                            <td style="text-align: center; background-color: #deeaf6; font-weight: bold;">1.</td>
                            <td style="font-weight: bold;">Logika Berpikir</td>
                            <td>Kemampuan untuk berpikir secara logis dan sistematis.</td>
                            {' '.join(f'<td style="{cell_style}">{x}</td>' for x in get_x_position(aspects["logika_berpikir"]))}
                        </tr>

                        <!-- Daya Analisa -->
                        <tr>
                            <td style="text-align: center; background-color: #deeaf6; font-weight: bold;">2.</td>
                            <td style="font-weight: bold;">Daya Analisa</td>
                            <td>Kemampuan untuk melihat permasalahan dan memahami hubungan sebab akibat permasalahan.</td>
                            {' '.join(f'<td style="{cell_style}">{x}</td>' for x in get_x_position(aspects["daya_analisa"]))}
                        </tr>

                        <!-- Kemampuan Numerikal -->
                        <tr>
                            <td style="text-align: center; background-color: #deeaf6; font-weight: bold;">3.</td>
                            <td style="font-weight: bold;">Kemampuan Numerikal</td>
                            <td>Kemampuan untuk berpikir praktis dalam memahami konsep angka dan hitungan.</td>
                            {' '.join(f'<td style="{cell_style}">{x}</td>' for x in get_x_position(aspects["kemampuan_numerik"]))}
                        </tr>

                        <!-- Kemampuan Verbal -->
                        <tr>
                            <td style="text-align: center; background-color: #deeaf6; font-weight: bold;">4.</td>
                            <td style="font-weight: bold;">Kemampuan Verbal</td>
                            <td>Kemampuan untuk memahami konsep dan pola dalam bentuk kata dan mengekspresikan gagasan secara verbal.</td>
                            {' '.join(f'<td style="{cell_style}">{x}</td>' for x in get_x_position(aspects["kemampuan_verbal"]))}
                        </tr>

                        <tr><td colspan="8" style="{section_header_style}">SIKAP DAN CARA KERJA</td></tr>

                        <!-- Orientasi Hasil -->
                        <tr>
                            <td style="text-align: center; background-color: #deeaf6; font-weight: bold;">5.</td>
                            <td style="font-weight: bold;">Orientasi Hasil</td>
                            <td>Kemampuan untuk mempertahankan komitmen untuk menyelesaikan tugas secara bertanggung jawab dan memperhatikan keterhubungan antara perencanaan dan hasil kerja.</td>
                            {' '.join(f'<td style="{cell_style}">{x}</td>' for x in get_x_position(aspects["orientasi_hasil"]))}
                        </tr>

                        <!-- Fleksibilitas -->
                        <tr>
                            <td style="text-align: center; background-color: #deeaf6; font-weight: bold;">6.</td>
                            <td style="font-weight: bold;">Fleksibilitas</td>
                            <td>Kemampuan untuk menyesuaikan diri dalam menghadapi permasalahan.</td>
                            {' '.join(f'<td style="{cell_style}">{x}</td>' for x in get_x_position(aspects["fleksibilitas"]))}
                        </tr>

                        <!-- Sistematika Kerja -->
                        <tr>
                            <td style="text-align: center; background-color: #deeaf6; font-weight: bold;">7.</td>
                            <td style="font-weight: bold;">Sistematika Kerja</td>
                            <td>Kemampuan untuk merencanakan hingga mengorganisasikan cara kerja dalam proses penyelesaian pekerjaannya.</td>
                            {' '.join(f'<td style="{cell_style}">{x}</td>' for x in get_x_position(aspects["sistematika_kerja"]))}
                        </tr>

                        <tr><td colspan="8" style="{section_header_style}">KEPRIBADIAN</td></tr>

                        <!-- Motivasi Berprestasi -->
                        <tr>
                            <td style="text-align: center; background-color: #deeaf6; font-weight: bold;">8.</td>
                            <td style="font-weight: bold;">Motivasi Berprestasi</td>
                            <td>Kemampuan untuk menunjukkan prestasi dan mencapai target.</td>
                            {' '.join(f'<td style="{cell_style}">{x}</td>' for x in get_x_position(aspects["motivasi_berprestasi"]))}
                        </tr>

                        <!-- Kerjasama -->
                        <tr>
                            <td style="text-align: center; background-color: #deeaf6; font-weight: bold;">9.</td>
                            <td style="font-weight: bold;">Kerjasama</td>
                            <td>Kemampuan untuk menjalin, membina dan mengoptimalkan hubungan kerja yang efektif demi tercapainya tujuan bersama.</td>
                            {' '.join(f'<td style="{cell_style}">{x}</td>' for x in get_x_position(aspects["kerjasama"]))}
                        </tr>

                        <!-- Keterampilan Interpersonal -->
                        <tr>
                            <td style="text-align: center; background-color: #deeaf6; font-weight: bold;">10.</td>
                            <td style="font-weight: bold;">Keterampilan Interpersonal</td>
                            <td>Kemampuan untuk menjalin hubungan sosial dan mampu memahami kebutuhan orang lain.</td>
                            {' '.join(f'<td style="{cell_style}">{x}</td>' for x in get_x_position(aspects["keterampilan_interpersonal"]))}
                        </tr>

                        <!-- Stabilitas Emosi -->
                        <tr>
                            <td style="text-align: center; background-color: #deeaf6; font-weight: bold;">11.</td>
                            <td style="font-weight: bold;">Stabilitas Emosi</td>
                            <td>Kemampuan untuk memahami dan mengontrol emosi.</td>
                            {' '.join(f'<td style="{cell_style}">{x}</td>' for x in get_x_position(aspects["stabilitas_emosi"]))}
                        </tr>

                        <tr><td colspan="8" style="{section_header_style}">KEMAMPUAN BELAJAR</td></tr>

                        <!-- Pengembangan Diri -->
                        <tr>
                            <td style="text-align: center; background-color: #deeaf6; font-weight: bold;">12.</td>
                            <td style="font-weight: bold;">Pengembangan Diri</td>
                            <td>Kemampuan untuk meningkatkan pengetahuan dan menyempurnakan keterampilan diri.</td>
                            {' '.join(f'<td style="{cell_style}">{x}</td>' for x in get_x_position(aspects["pengembangan_diri"]))}
                        </tr>

                        <!-- Mengelola Perubahan -->
                        <tr>
                            <td style="text-align: center; background-color: #deeaf6; font-weight: bold;">13.</td>
                            <td style="font-weight: bold;">Mengelola Perubahan</td>
                            <td>Kemampuan dalam menyesuaikan diri dengan situasi yang baru.</td>
                            {' '.join(f'<td style="{cell_style}">{x}</td>' for x in get_x_position(aspects["mengelola_perubahan"]))}
                        </tr>

                        <!-- Legend -->
                        <tr style="border-top: 1px solid black;">
                            <td colspan="8" style="text-align: center; padding: 2px; font-family: Arial; font-size: 11px; background-color: #deeaf6;">
                                <div style="display: inline-block; width: 100%; font-weight: bold;">
                                    T : Tinggi&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;
                                    B : Baik&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;
                                    C : Cukup&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;
                                    K : Kurang&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;
                                    R : Rendah
                                </div>
                            </td>
                        </tr>
                    </table>
                """

            # Get the gender value from the selected row
            jk_col = self.get_column_index("JK")
            jk_val = self.table.item(selected_row, jk_col)
            gender_prefix = "Sdr."

            if jk_val:
                gender_text = jk_val.text().strip().upper()
                if gender_text == "P":
                    gender_prefix = "Sdri."

            # Ensure the Excel file path is set
            excel_file_path = self.excel_file_path

            # Load the workbook and select the appropriate sheet
            wb = openpyxl.load_workbook(excel_file_path, data_only=True)
            sheet3 = wb.worksheets[2]  # Ensure this is the correct sheet index

            def get_development_text(aspect_value, r_text, k_text):
                # Return appropriate text based on R or K value
                if aspect_value == "R":
                    return r_text
                elif aspect_value == "K":
                    return k_text
                return ""

            # Map each specified column to its corresponding development text
            column_mappings = {
                "Logika Berpikir 1": ("F3", "F4"),
                "Daya Analisa 3": ("F5", "F6"),
                "Kemampuan Numerik 5": ("F7", "F8"),
                "Kemampuan Verbal 2 dam 4": ("F9", "F10"),
                "Orientasi Hasil/ N G": ("F11", "F12"),
                "Fleksibilitas/ T V": ("F13", "F14"),
                "Sistematika Kerja/ C D R": ("F15", "F16"),
                "Motivasi Berprestasi/ A": ("F17", "F18"),
                "Kerjasama/ P I": ("F19", "F20"),
                "Keterampilan Interpersonal/ B S": ("F21", "F22"),
                "Stabilitas Emosi/ E PHQ": ("F23", "F24"),
                "Pegembangan Diri/ W": ("F25", "F26"),
                "Mengelola Perubahan/ Z K": ("F27", "F28")
            }

            # Get development text only for columns with R or K value
            development_texts = []
            for column, (r_cell, k_cell) in column_mappings.items():
                column_value = row_data.get(column, "")
                if column_value in ["R", "K"]:
                    text = get_development_text(column_value, sheet3[r_cell].value or "", sheet3[k_cell].value or "")
                    if text:
                        development_texts.append(text)

            # Build final development text
            development_text = f"{gender_prefix} {nama} "
            
            if development_texts:
                # Join only the texts from columns that had R or K values
                development_text += " ".join(development_texts)
            else:
                # If no R or K values found, use a default message
                development_text += "masih membutuhkan pengembangan dalam mengidentifikasi pola logis, menarik kesimpulan, dan memperdalami penguasaan hubungan sebab akibat menjadi lebih baik dalam konsep matematis dan butuh memperdalam penggunaan bahasa lebih lanjut. Kemudian, butuh kefokusan agar lebih mudah dalam beradaptasi, dan butuh kontribusi lebih, dan koordinasi dengan kelompok agar mencapai tujuan bersama. Serta mudah diri dengan pada pada hal-hal baru."

            # Get the PHQ value from the selected row
            phq_col = self.get_column_index("PHQ")
            phq_val = self.table.item(selected_row, phq_col)

            screening_categories = {
                "Tahapan Normal": "Individu menunjukkan adaptasi gejala gangguan mental yang mengganggu fungsi sehari-hari",
                "Cenderung Stress dalam Tekanan": "Dalam situasi yg menimbulkan tekanan dapat berdampak pada kondisi individu & respon emosional yg ditampilkan",
                "Gejala Gangguan": "Individu menunjukkan gejala-gejala gangguan yang dapat mengganggu fungsi sehari-hari"
            }
            selected_category = ""
            if phq_val:
                try:
                    phq_value_numeric = int(phq_val.text().strip())
                    if 0 <= phq_value_numeric <= 9:
                        selected_category = "Tahapan Normal"
                    elif 10 <= phq_value_numeric <= 19:
                        selected_category = "Cenderung Stress dalam Tekanan"
                    elif phq_value_numeric >= 20:
                        selected_category = "Gejala Gangguan"
                except ValueError:
                    selected_category = "Unknown"

            def generate_conclusion_text(aspect, category, sheet, option):
                # Define the cell mappings for each aspect and category
                cell_mappings = {
                    "Logika Berpikir 1": {
                        "R": {"F": "F30", "G": "G30", "H": "H30"},
                        "K": {"F": "F31", "G": "G31", "H": "H31"},
                        "C": {"F": "F32", "G": "G32", "H": "H32"},
                        "B": {"F": "F33", "G": "G33", "H": "H33"},
                        "T": {"F": "F34", "G": "G34", "H": "H34"}
                    },
                    "Daya Analisa 3": {
                        "R": {"F": "F35", "G": "G35", "H": "H35"},
                        "K": {"F": "F36", "G": "G36", "H": "H36"},
                        "C": {"F": "F37", "G": "G37", "H": "H37"},
                        "B": {"F": "F38", "G": "G38", "H": "H38"},
                        "T": {"F": "F39", "G": "G39", "H": "H39"}
                    },
                    "Kemampuan Numerik 5": {
                        "R": {"F": "F40", "G": "G40", "H": "H40"},
                        "K": {"F": "F41", "G": "G41", "H": "H41"},
                        "C": {"F": "F42", "G": "G42", "H": "H42"},
                        "B": {"F": "F43", "G": "G43", "H": "H43"},
                        "T": {"F": "F44", "G": "G44", "H": "H44"}
                    },
                    "Kemampuan Verbal 2 dam 4": {
                        "R": {"F": "F45", "G": "G45", "H": "H45"},
                        "K": {"F": "F46", "G": "G46", "H": "H46"},
                        "C": {"F": "F47", "G": "G47", "H": "H47"},
                        "B": {"F": "F48", "G": "G48", "H": "H48"},
                        "T": {"F": "F49", "G": "G49", "H": "H49"}
                    },
                    "Orientasi Hasil/ N G": {
                        "R": {"F": "F50", "G": "G50", "H": "H50"},
                        "K": {"F": "F51", "G": "G51", "H": "H51"},
                        "C": {"F": "F52", "G": "G52", "H": "H52"},
                        "B": {"F": "F53", "G": "G53", "H": "H53"},
                        "T": {"F": "F54", "G": "G54", "H": "H54"}
                    },
                    "Fleksibilitas/ T V": {
                        "R": {"F": "F55", "G": "G55", "H": "H55"},
                        "K": {"F": "F56", "G": "G56", "H": "H56"},
                        "C": {"F": "F57", "G": "G57", "H": "H57"},
                        "B": {"F": "F58", "G": "G58", "H": "H58"},
                        "T": {"F": "F59", "G": "G59", "H": "H59"}
                    },
                    "Sistematika Kerja/ C D R": {
                        "R": {"F": "F60", "G": "G60", "H": "H60"},
                        "K": {"F": "F61", "G": "G61", "H": "H61"},
                        "C": {"F": "F62", "G": "G62", "H": "H62"},
                        "B": {"F": "F63", "G": "G63", "H": "H63"},
                        "T": {"F": "F64", "G": "G64", "H": "H64"}
                    },
                    "Motivasi Berprestasi/ A": {
                        "R": {"F": "F65", "G": "G65", "H": "H65"},
                        "K": {"F": "F66", "G": "G66", "H": "H66"},
                        "C": {"F": "F67", "G": "G67", "H": "H67"},
                        "B": {"F": "F68", "G": "G68", "H": "H68"},
                        "T": {"F": "F69", "G": "G69", "H": "H69"}
                    },
                    "Kerjasama/ P I": {
                        "R": {"F": "F70", "G": "G70", "H": "H70"},
                        "K": {"F": "F71", "G": "G71", "H": "H71"},
                        "C": {"F": "F72", "G": "G72", "H": "H72"},
                        "B": {"F": "F73", "G": "G73", "H": "H73"},
                        "T": {"F": "F74", "G": "G74", "H": "H74"}
                    },
                    "Keterampilan Interpersonal/ B S": {
                        "R": {"F": "F75", "G": "G75", "H": "H75"},
                        "K": {"F": "F76", "G": "G76", "H": "H76"},
                        "C": {"F": "F77", "G": "G77", "H": "H77"},
                        "B": {"F": "F78", "G": "G78", "H": "H78"},
                        "T": {"F": "F79", "G": "G79", "H": "H79"}
                    },
                    "Stabilitas Emosi/ E PHQ": {
                        "R": {"F": "F80", "G": "G80", "H": "H80"},
                        "K": {"F": "F81", "G": "G81", "H": "H81"},
                        "C": {"F": "F82", "G": "G82", "H": "H82"},
                        "B": {"F": "F83", "G": "G83", "H": "H83"},
                        "T": {"F": "F84", "G": "G84", "H": "H84"}
                    },
                    "Pegembangan Diri/ W": {
                        "R": {"F": "F85", "G": "G85", "H": "H85"},
                        "K": {"F": "F86", "G": "G86", "H": "H86"},
                        "C": {"F": "F87", "G": "G87", "H": "H87"},
                        "B": {"F": "F88", "G": "G88", "H": "H88"},
                        "T": {"F": "F89", "G": "G89", "H": "H89"}
                    },
                    "Mengelola Perubahan/ Z K": {
                        "R": {"F": "F90", "G": "G90", "H": "H90"},
                        "K": {"F": "F91", "G": "G91", "H": "H91"},
                        "C": {"F": "F92", "G": "G92", "H": "H92"},
                        "B": {"F": "F93", "G": "G93", "H": "H93"},
                        "T": {"F": "F94", "G": "G94", "H": "H94"}
                    }
                }
                
                # Get the cell reference for the given aspect, category, and option
                cell_ref = cell_mappings.get(aspect, {}).get(category, {}).get(option, "")
                
                # Retrieve the text from the specified cell
                if cell_ref:
                    cell_value = sheet[cell_ref].value
                    return cell_value if cell_value is not None else f"Default text for {aspect} and category {category}"
                else:
                    return f"Default text for {aspect}"

            # Define aspect categories
            aspect_categories = {
                "KEMAMPUAN INTELEKTUAL": ["Logika Berpikir 1", "Daya Analisa 3", "Kemampuan Numerik 5", "Kemampuan Verbal 2 dam 4"],
                "SIKAP DAN CARA KERJA": ["Orientasi Hasil/ N G", "Fleksibilitas/ T V", "Sistematika Kerja/ C D R"],
                "KEPRIBADIAN": ["Motivasi Berprestasi/ A", "Kerjasama/ P I", "Keterampilan Interpersonal/ B S", "Stabilitas Emosi/ E PHQ"],
                "KEMAMPUAN BELAJAR": ["Pegembangan Diri/ W", "Mengelola Perubahan/ Z K"]
            }

            # Generate conclusion text for each aspect
            conclusion_texts = {category: [] for category in aspect_categories}

            # Predefined option choice (set to "F", "G", or "H")
            # Randomly select an option for variety
            option = random.choice(["F", "G", "H"])

            # Function to choose an option based on the predefined choice
            def choose_option(aspect, current_option):
                # Always return the predefined option
                return current_option

            # Get IQ value first
            iq_val = self.table.item(selected_row, self.get_column_index("IQ "))
            iq_value = float(iq_val.text()) if iq_val and iq_val.text().strip() else 0

            # Initialize counters for recommendation criteria
            intellectual_k_count = 0
            work_attitude_k_count = 0 
            personality_k_count = 0

            # Iterate over aspects and use the chosen option consistently
            for aspect in column_mappings.keys():
                category = row_data.get(aspect, "")
                text = generate_conclusion_text(aspect, category, sheet3, option)
                
                # Count K values for each category
                if category == "K":
                    if aspect in aspect_categories["KEMAMPUAN INTELEKTUAL"]:
                        intellectual_k_count += 1
                    elif aspect in aspect_categories["SIKAP DAN CARA KERJA"]:
                        work_attitude_k_count += 1
                    elif aspect in aspect_categories["KEPRIBADIAN"]:
                        personality_k_count += 1
                        
                for category_name, aspects in aspect_categories.items():
                    if aspect in aspects:
                        conclusion_texts[category_name].append(text)

            # Convert aspects dictionary values to lists with default empty strings
            intellectual_scores = []
            for key in ["daya_analisa", "kemampuan_numerik", "kemampuan_verbal"]:
                intellectual_scores.append(aspects[key] if key in aspects else "")

            work_attitude_scores = []
            for key in ["sistematika_kerja", "fleksibilitas"]:
                work_attitude_scores.append(aspects[key] if key in aspects else "")

            personality_scores = []
            for key in ["inisiatif", "kerjasama", "keterampilan_interpersonal", "stabilitas_emosi"]:
                personality_scores.append(aspects[key] if key in aspects else "")

            # Determine recommendation based on criteria
            def determine_recommendation():
                if iq_value >= 90 and intellectual_k_count <= 2 and work_attitude_k_count <= 1 and personality_k_count <= 1:
                    return "LAYAK DIREKOMENDASIKAN"
                elif 86 <= iq_value < 90 and intellectual_k_count <= 3 and work_attitude_k_count <= 1 and personality_k_count <= 2:
                    return "LAYAK DIPERTIMBANGKAN"
                else:
                    return "TIDAK DISARANKAN"

            # Get the overall recommendation
            overall_recommendation = determine_recommendation()
            
            # Add page break and second page content
            html_content += f"""
                <div class="page-break"></div>
                <div class="page" style="font-family: Arial; width: 21cm; height: 29.7cm; margin: 0 auto; padding: 1cm; position: relative;">
                    <div style="position: relative; display: flex; justify-content: space-between; align-items: center; margin-bottom: 40px;">
                        <div style="width: 150px;">
                            <img src="{logo_data_url}" alt="Logo" style="width: 100%; height: auto;">
                        </div>
                        <div style="text-align: center; flex-grow: 1;">
                        <div style="font-size: 15px; font-weight: bold; color: #1f4e79;">HASIL PEMERIKSAAN PSIKOLOGIS</div>
                        <div style="font-size: 17px; font-weight: bold; color: #1f4e79;">(Asesmen Intelegensi, Kepribadian dan Minat)</div>
                        </div>
                        <div style="text-align: right; font-size: 12px;">
                            <div style="font-weight: bold; color: #1f4e79;">RAHASIA</div>
                            <div style="color: #1f4e79;">No. {row_data.get('No', '')}{row_data.get('No Tes', '')}</div>
                        </div>
                    </div>

                    <table class="psikogram" style="width: 100%; border-collapse: collapse; margin-top: 10px; font-family: Arial, sans-serif;">
                        <tr>
                            <th colspan="2" style="text-align: center; padding: 8px; background-color: #fbe4d5; border: 1px solid black; font-size: 16px;">KESIMPULAN</th>
                        </tr>
                        <tr>
                            <td style="width: 20%; padding: 8px; vertical-align: top; border: 1px solid black; font-weight: bold; font-size: 12px; ">KEMAMPUAN INTELEKTUAL</td>
                            <td style="width: 80%; padding: 8px; text-align: justify; border: 1px solid black; font-size: 14px;">
                                {row_data.get('Intelegensi Umum.1', '')}
                                Berdasarkan pemeriksaan kemampuan intelektual, diketahui bahwa {gender_prefix} {nama} memiliki {' '.join(conclusion_texts["KEMAMPUAN INTELEKTUAL"])}
                            </td>
                        </tr>
                        <tr>
                            <td style="padding: 8px; vertical-align: top; border: 1px solid black; font-weight: bold; font-size: 12px; ">SIKAP DAN CARA KERJA</td>
                            <td style="padding: 8px; text-align: justify; border: 1px solid black; font-size: 14px;">
                                {row_data.get('Sistematika Kerja/ cd.1', '')}
                                Berdasarkan pemeriksaan sikap dan cara kerja, diketahui bahwa {gender_prefix} {nama} menunjukkan {' '.join(conclusion_texts["SIKAP DAN CARA KERJA"])}
                            </td>
                        </tr>
                        <tr>
                            <td style="padding: 8px; vertical-align: top; border: 1px solid black; font-weight: bold ; font-size: 12px; ">KEPRIBADIAN</td>
                            <td style="padding: 8px; text-align: justify; border: 1px solid black; font-size: 14px;">
                                {row_data.get('Stabilitas Emosi / E.1', '')}
                                Berdasarkan pemeriksaan kepribadian, diketahui bahwa {gender_prefix} {nama} memiliki {' '.join(conclusion_texts["KEPRIBADIAN"])}
                            </td>
                        </tr>
                        <tr>
                            <td style="padding: 8px; vertical-align: top; border: 1px solid black; font-weight: bold; font-size: 12px;">KEMAMPUAN BELAJAR</td>
                            <td style="padding: 8px; text-align: justify; border: 1px solid black; font-size: 14px;">
                                {row_data.get('Fleksibilitas', '')}
                                Berdasarkan pemeriksaan kemampuan belajar, diketahui bahwa {gender_prefix} {nama} menunjukkan {' '.join(conclusion_texts["KEMAMPUAN BELAJAR"])}
                            </td>
                        </tr>
                    </table>
                    
                    <table class="psikogram" style="width: 100%; border-collapse: collapse; margin-top: 20px; font-family: Arial, sans-serif;">
                        <tr>
                            <th colspan="2" style="text-align: center; padding: 12px; background-color: #fbe4d5; border: 1px solid black; font-size: 14px;">PENGEMBANGAN</th>
                        </tr>
                        <tr>
                            <td colspan="2" style="padding: 12px; text-align: justify; border: 1px solid black; font-size: 14px;">
                                {development_text}
                            </td>
                        </tr>
                    </table>

                    <table class="psikogram" style="width: 100%; border-collapse: collapse; margin-top: 20px; font-family: Arial, sans-serif;">
                        <tr>
                            <th colspan="3" style="text-align: center; padding: 8px; background-color: #fbe4d5; border: 1px solid black; font-size: 14px;">Kategori Hasil Screening</th>
                        </tr>
                        <tr>
                            <td style="width: 10%; text-align: center; border: 1px solid black; padding: 10px; font-size: 14px;">{"X" if selected_category == "Tahapan Normal" else ""}</td>
                            <td style="width: 30%; text-align: center; padding: 10px; border: 1px solid black; font-size: 14px;">Tahapan Normal</td>
                            <td style="padding: 10px; border: 1px solid black; font-size: 14px;">Individu tidak menunjukkan adanya gejala gangguan mental yang mengganggu fungsi sehari-hari.</td>
                        </tr>
                        <tr>
                            <td style="text-align: center; border: 1px solid black; padding: 10px; font-size: 14px;">{"X" if selected_category == "Cenderung Stress dalam Tekanan" else ""}</td>
                            <td style="text-align: center; padding: 10px; border: 1px solid black; font-size: 14px;">Kecenderungan Stress dalam Tekanan</td>
                            <td style="padding: 10px; border: 1px solid black; font-size: 14px;">Dalam situasi yg menimbulkan tekanan dapat berdampak pada kondisi mental & respon emosional yg ditampilkan.</td>
                        </tr>
                        <tr>
                            <td style="text-align: center; border: 1px solid black; padding: 10px; font-size: 14px;">{"X" if selected_category == "Gejala Gangguan" else ""}</td>
                            <td style="text-align: center; padding: 10px; border: 1px solid black; font-size: 14px;">Gangguan</td>
                            <td style="padding: 10px; border: 1px solid black; font-size: 14px;">Individu menunjukkan gejala gangguan mental yang dapat mengganggu fungsi sehari-hari.</td>
                        </tr>
                    </table>
                    
                    <table class="psikogram" style="width: 40%; border-collapse: collapse; margin-top: 20px; font-family: Arial, sans-serif;">
                        <tr>
                            <th colspan="2" style="text-align: center; padding: 8px; background-color: #fbe4d5; border: 1px solid black;">Kesimpulan Keseluruhan</th>
                        </tr>
                        <tr>
                            <td style="width: 10%; text-align: center; border: 1px solid black; padding: 8px;">{"X" if overall_recommendation == "LAYAK DIREKOMENDASIKAN" else ""}</td>
                            <td style="padding: 8px; border: 1px solid black;">LAYAK DIREKOMENDASIKAN</td>
                        </tr>
                        <tr>
                            <td style="width: 10%; text-align: center; border: 1px solid black; padding: 8px;">{"X" if overall_recommendation == "LAYAK DIPERTIMBANGKAN" else ""}</td>
                            <td style="padding: 8px; border: 1px solid black;">LAYAK DIPERTIMBANGKAN</td>
                        </tr>
                        <tr>
                            <td style="width: 10%; text-align: center; border: 1px solid black; padding: 8px;">{"X" if overall_recommendation == "TIDAK DISARANKAN" else ""}</td>
                            <td style="padding: 8px; border: 1px solid black;">TIDAK DISARANKAN</td>
                        </tr>
                    </table>
                </div>
            """

            # Add page break and third page content
            html_content += f"""
                <div class="page-break"></div>
                <div class="page" style="padding: 1cm; font-family: Arial;">
                    <div style="display: flex; align-items: center; margin-bottom: 20px;">
                        <img src="{logo_data_url}" alt="Logo" style="width: 150px; height: auto; margin-right: 20px;">
                        <div style="flex-grow: 1; text-align: center;">
                        <div style="font-size: 14px; font-weight: bold; color: #1f4e79;">HASIL PEMERIKSAAN PSIKOLOGIS</div>
                        <div style="font-size: 16px; font-weight: bold; color: #1f4e79;">(Asesmen Intelegensi, Kepribadian dan Minat)</div>
                        </div>
                        <div style="text-align: right; font-size: 12px;">
                            <div style="font-weight: bold; color: #1f4e79;">RAHASIA</div>
                            <div style="color: #1f4e79;">{row_data.get('No', '')}{row_data.get('No Tes', '')}</div>
                        </div>
                    </div>

                    <div style="margin-bottom: 20px;">
                        <div style="margin-bottom: 15px;">
                            <div>
                                <span style="display: inline-block; width: 120px; font-size: 14px;">Tanggal</span>
                                <span style="font-size: 14px;">: {datetime.now().strftime("%d %B %Y")}</span>
                            </div>
                            <div style="font-style: italic; font-size: 12px; color: #666;">Date</div>
                        </div>
                        
                        <div style="margin-bottom: 15px;">
                            <div>
                                <span style="display: inline-block; width: 120px; font-size: 14px;">Tanda Tangan</span>
                            </div>
                            <div style="font-style: italic; font-size: 12px; color: #666;">Signature</div>
                        </div>
                        
                        <div style="margin-bottom: 15px;">
                            <div>
                                <span style="display: inline-block; width: 120px; font-size: 14px;">Nama Psikolog</span>
                                <span style="font-size: 14px;">: Chitra Ananda Mulia, M.Psi., Psikolog</span>
                            </div>
                            <div style="font-style: italic; font-size: 12px; color: #666;">Psychologist Name</div>
                        </div>
                        
                        <div style="margin-bottom: 15px;">
                            <div>
                                <span style="display: inline-block; width: 120px; font-size: 14px;">Nomor STR/SIK</span>
                                <span style="font-size: 14px;">:</span>
                            </div>
                            <div style="font-style: italic; font-size: 12px; color: #666;">Registration Number</div>
                        </div>
                        
                        <div style="margin-bottom: 15px;">
                            <div>
                                <span style="display: inline-block; width: 120px; font-size: 14px;">Nomor SIPP/SIPPK</span>
                                <span style="font-size: 14px;">: 1564-19-2-2</span>
                            </div>
                            <div style="font-style: italic; font-size: 12px; color: #666;">Licence Number</div>
                        </div>
                    </div>
                </div>
            """

            # Close tables and add footer
            html_content += """
                </table>
                <div class="footer">
                    Laporan ini bersifat confidential dan diketahui oleh Psikolog
                </div>
            </body>
            </html>
            """

            # Create and show preview dialog
            preview_dialog = QDialog(self)
            preview_dialog.setWindowTitle("Preview PDF")
            
            # Get screen size
            screen = QApplication.primaryScreen()
            screen_size = screen.availableGeometry()
            
            # Set dialog size to 90% of screen size
            dialog_width = int(screen_size.width() * 0.9)
            dialog_height = int(screen_size.height() * 0.9)
            preview_dialog.setFixedSize(dialog_width, dialog_height)
            
            # Center the dialog on screen
            preview_dialog.move(
                (screen_size.width() - dialog_width) // 2,
                (screen_size.height() - dialog_height) // 2
            )
            
            preview_dialog.setWindowFlags(Qt.Window | Qt.WindowMaximizeButtonHint | Qt.WindowCloseButtonHint)
            
            # Create main vertical layout
            main_layout = QVBoxLayout(preview_dialog)
            main_layout.setContentsMargins(10, 10, 10, 10)
            main_layout.setSpacing(10)
            
            # Create horizontal layout for preview pages
            preview_layout = QHBoxLayout()
            
            # Create web view for page 1
            web_view1 = QWebEngineView(preview_dialog)
            web_view1.setZoomFactor(0.6)
            web_view1.setFixedWidth(int(dialog_width * 0.3))  # Adjust width to 30% for 3 pages

            # Split HTML content at page break
            pages = html_content.split('<div class="page-break"></div>')
            web_view1.setHtml(pages[0])
            preview_layout.addWidget(web_view1)
            
            # Create web view for page 2
            web_view2 = QWebEngineView(preview_dialog)
            web_view2.setZoomFactor(0.55)
            web_view2.setFixedWidth(int(dialog_width * 0.3))  # Adjust width to 30% for 3 pages
            if len(pages) > 1:
                web_view2.setHtml(pages[1])
            preview_layout.addWidget(web_view2)

            # Create web view for page 3
            web_view3 = QWebEngineView(preview_dialog)
            web_view3.setZoomFactor(0.7)
            web_view3.setFixedWidth(int(dialog_width * 0.3))  # Adjust width to 30% for 3 pages
            if len(pages) > 2:
                web_view3.setHtml(pages[2])
            preview_layout.addWidget(web_view3)
            
            # Add preview layout to main layout
            main_layout.addLayout(preview_layout)
            
            # Create button layout for centering
            button_layout = QHBoxLayout()
            button_layout.addStretch()
            
            # Add save PDF button
            save_button = QPushButton("Save PDF", preview_dialog)
            save_button.setFixedHeight(30)
            save_button.setFixedWidth(200)  # Set fixed width for button
            save_button.clicked.connect(lambda: self.save_as_pdf(html_content))
            button_layout.addWidget(save_button)
            
            button_layout.addStretch()
            main_layout.addLayout(button_layout)
            
            preview_dialog.exec_()
            
        except Exception as e:
            print(f"Error saat preview: {e}")
            
    def print_pdf(self, html_content):
        try:
            # Create printer with A4 settings
            printer = QPrinter(QPrinter.HighResolution)
            printer.setPageSize(QPageSize(QPageSize.A4))
            # Set margins using individual float values
            printer.setPageMargins(10, 10, 10, 10, QPrinter.Millimeter)
            
            # Create web view and load content
            web_view = QWebEngineView()
            web_view.setHtml(html_content)
            
            # Wait for page load
            loop = QEventLoop()
            web_view.loadFinished.connect(loop.quit)
            loop.exec_()
            
            # Show print dialog
            print_dialog = QPrintDialog(printer, self)
            if print_dialog.exec_() == QPrintDialog.Accepted:
                def print_finished(success):
                    if success:
                        QMessageBox.information(self, "Success", "Document printed successfully")
                    else:
                        QMessageBox.warning(self, "Warning", "Print job failed")
                
                # Direct print without preview
                web_view.page().print(printer, print_finished)
                
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Error printing document: {e}")
            print(f"Print error: {e}")

    def save_as_pdf(self, html_content):
        try:
            # Get save file name from dialog
            file_name, _ = QFileDialog.getSaveFileName(
                self,
                "Save PDF", 
                "",
                "PDF Files (*.pdf)"
            )
            
            if not file_name:
                return  # User canceled
                
            # Ensure .pdf extension
            if not file_name.lower().endswith('.pdf'):
                file_name += '.pdf'

            # Create a printer with PDF output and optimize settings
            printer = QPrinter(QPrinter.HighResolution)
            printer.setOutputFormat(QPrinter.PdfFormat)
            printer.setOutputFileName(file_name)
            printer.setPageSize(QPageSize(QPageSize.A4))
            printer.setPageMargins(10, 10, 10, 10, QPrinter.Millimeter)
            
            # Set PDF optimization options
            printer.setResolution(300) # Optimal resolution for clear text
            printer.setPdfVersion(QPrinter.PdfVersion_1_6) # Use optimized PDF version

            # Create web view for PDF generation
            web_view = QWebEngineView()
            
            # Add styles with optimized settings
            styled_content = html_content.replace('</head>',
            '''
            <style>
                @page {
                    size: A4;
                    margin: 1cm;
                }
                @media print {
                    body {
                        width: 210mm;
                        height: 297mm;
                        margin: 0;
                        padding: 1cm;
                    }
                    .page {
                        page-break-after: always;
                    }
                    .page:last-child {
                        page-break-after: avoid;
                    }
                    img {
                        -webkit-optimize-contrast: true;
                        image-rendering: optimizeQuality;
                        max-width: 100%;
                        height: auto;
                    }
                }
                body {
                    margin: 0;
                    padding: 1cm;
                    width: 210mm;
                    height: 297mm;
                    font-family: Arial, sans-serif;
                    font-size: 11px;
                    position: relative;
                    text-rendering: optimizeLegibility;
                }
                .header {
                    text-align: center;
                    margin-bottom: 15px;
                }
                .header .title {
                    font-size: 14px;
                    font-weight: bold;
                    margin-bottom: 5px;
                }
                .info-table {
                    width: 100%;
                    margin-bottom: 15px;
                    border-spacing: 0;
                }
                .info-table td {
                    padding: 3px;
                    vertical-align: top;
                }
                table {
                    border-collapse: collapse;
                    width: 100%;
                }
                .psikogram {
                    margin-top: 15px;
                }
                .psikogram th, .psikogram td {
                    border: 1px solid black;
                    padding: 4px;
                    font-size: 11px;
                }
                .psikogram th {
                    background-color: #f2f2f2;
                    text-align: center;
                }
                .category-header {
                    background-color: #f8d7da !important;
                    text-align: center;
                    font-weight: bold;
                }
                .footer {
                    position: fixed;
                    bottom: 1cm;
                    left: 1cm;
                    right: 1cm;
                    text-align: center;
                    font-style: italic;
                    font-size: 10px;
                    background: white;
                    padding: 5px;
                }
                .legend-row td {
                    text-align: center;
                    padding: 2px;
                    font-size: 11px;
                    border: none;
                }
                .page-footer {
                    position: fixed;
                    bottom: 1cm;
                    left: 1cm;
                    right: 1cm;
                    text-align: center;
                    font-style: italic;
                    font-size: 10px;
                    background: white;
                    padding: 5px;
                }
            </style>
            </head>
            ''')

            # Load content into web view and handle PDF generation
            web_view.loadFinished.connect(lambda: self.handle_pdf_generation(web_view, printer))
            web_view.setHtml(styled_content)

            # Keep reference to web view to prevent garbage collection
            self._temp_web_view = web_view

        except Exception as e:
            QMessageBox.critical(self, "Error", f"Error saving PDF: {e}")
            print(f"Error saving PDF: {e}")

    def handle_pdf_generation(self, web_view, printer):
        def print_finished(success):
            if success:
                QMessageBox.information(self, "Success", "PDF saved successfully!")
            else:
                QMessageBox.critical(self, "Error", "Failed to save PDF")

        web_view.page().print(printer, print_finished)

    def save_pdf_file(self, pdf_data, file_name):
        try:
            with open(file_name, 'wb') as f:
                f.write(pdf_data)
                f.flush()
                os.fsync(f.fileno())
        except Exception as e:
            print(f"Error writing PDF file: {e}")

    def apply_psikogram_formulas(self, sheet, row_idx):
        """
        Menerapkan rumus-rumus psikogram untuk baris tertentu
        """
        try:
            # Logika Berpikir 1
            logika_cell = sheet.cell(row=row_idx, column=45)
            logika_cell.value = f'=IF(L{row_idx}<80,"R",IF(L{row_idx}<100,"K",IF(L{row_idx}<120,"C",IF(L{row_idx}<140,"B","T"))))'

            # Daya Analisa 3
            analisa_cell = sheet.cell(row=row_idx, column=46)
            analisa_cell.value = f'=IF(M{row_idx}<80,"R",IF(M{row_idx}<100,"K",IF(M{row_idx}<120,"C",IF(M{row_idx}<140,"B","T"))))'

            # Kemampuan Verbal 2 dam 4
            verbal_cell = sheet.cell(row=row_idx, column=47)
            verbal_cell.value = f'=IF(O{row_idx}<80,"R",IF(O{row_idx}<100,"K",IF(O{row_idx}<120,"C",IF(O{row_idx}<140,"B","T"))))'

            # Kemampuan Numerik 5
            numerik_cell = sheet.cell(row=row_idx, column=48)
            numerik_cell.value = f'=IF(N{row_idx}<80,"R",IF(N{row_idx}<100,"K",IF(N{row_idx}<120,"C",IF(N{row_idx}<140,"B","T"))))'

            # Sistematika Kerja/ C D R
            sistematika_cell = sheet.cell(row=row_idx, column=49)
            sistematika_cell.value = f'=IFS(AO{row_idx}<2,"R",AO{row_idx}<4,"K",AO{row_idx}<6,"C",AO{row_idx}<9,"B",AO{row_idx}=9,"T")'

            # Orientasi Hasil/ N G
            orientasi_cell = sheet.cell(row=row_idx, column=50)
            orientasi_cell.value = f'=IFS(AN{row_idx}<2,"R",AN{row_idx}<4,"K",AN{row_idx}<6,"C",AN{row_idx}<9,"B",AN{row_idx}=9,"T")'

            # Fleksibilitas/ T V
            fleksibilitas_cell = sheet.cell(row=row_idx, column=51)
            fleksibilitas_cell.value = f'=IFS(AP{row_idx}<2,"R",AP{row_idx}<4,"K",AP{row_idx}<6,"C",AP{row_idx}<9,"B",AP{row_idx}=9,"T")'

            # Motivasi Berprestasi/ A
            motivasi_cell = sheet.cell(row=row_idx, column=52)
            motivasi_cell.value = f'=IFS(U{row_idx}<2,"R",U{row_idx}<4,"K",U{row_idx}<6,"C",U{row_idx}<9,"B",U{row_idx}=9,"T")'

            # Kerjasama/ P I
            kerjasama_cell = sheet.cell(row=row_idx, column=53)
            kerjasama_cell.value = f'=IFS(AQ{row_idx}<2,"R",AQ{row_idx}<4,"K",AQ{row_idx}<6,"C",AQ{row_idx}<9,"B",AQ{row_idx}=9,"T")'

            # Keterampilan Interpersonal/ B S
            interpersonal_cell = sheet.cell(row=row_idx, column=54)
            interpersonal_cell.value = f'=IFS(AR{row_idx}<2,"R",AR{row_idx}<4,"K",AR{row_idx}<6,"C",AR{row_idx}<9,"B",AR{row_idx}=9,"T")'

            # Stabilitas Emosi/ E PHQ
            emosi_cell = sheet.cell(row=row_idx, column=55)
            emosi_cell.value = f'=IFS(AJ{row_idx}<2,"R",AJ{row_idx}<4,"K",AJ{row_idx}<6,"C",AJ{row_idx}<9,"B",AJ{row_idx}=9,"T")'

            # Pegembangan Diri/ W
            pengembangan_cell = sheet.cell(row=row_idx, column=56)
            pengembangan_cell.value = f'=IFS(AM{row_idx}<2,"R",AM{row_idx}<4,"K",AM{row_idx}<6,"C",AM{row_idx}<9,"B",AM{row_idx}=9,"T")'

            # Mengelola Perubahan/ Z K
            perubahan_cell = sheet.cell(row=row_idx, column=57)
            perubahan_cell.value = f'=IFS(AS{row_idx}<2,"R",AS{row_idx}<4,"K",AS{row_idx}<6,"C",AS{row_idx}<9,"B",AS{row_idx}=9,"T")'

            # Kolom 59-71: Formula untuk kolom .1
            try:
                # Logika Berpikir 1.1 (kolom 59)
                logika_cell = sheet.cell(row=row_idx, column=58)
                logika_cell.value = f'=IF(AT{row_idx}="R",Sheet3!$F$30,IF(AT{row_idx}="K",Sheet3!$F$31,IF(AT{row_idx}="C",Sheet3!$F$32,IF(AT{row_idx}="B",Sheet3!$F$33,IF(AT{row_idx}="T",Sheet3!$F$34)))))'

                # Daya Analisa 3.1 (kolom 59)
                analisa_cell = sheet.cell(row=row_idx, column=59)
                analisa_cell.value = f'=IF(AU{row_idx}="R",Sheet3!$F$35,IF(AU{row_idx}="K",Sheet3!$F$36,IF(AU{row_idx}="C",Sheet3!$F$37,IF(AU{row_idx}="B",Sheet3!$F$38,IF(AU{row_idx}="T",Sheet3!$F$389)))))'

                # Kemampuan Verbal 2 dam 4.1 (kolom 60)
                verbal_cell = sheet.cell(row=row_idx, column=60)
                verbal_cell.value = f'=IF(AV{row_idx}="R",Sheet3!$F$45,IF(AV{row_idx}="K",Sheet3!$F$46,IF(AV{row_idx}="C",Sheet3!$F$47,IF(AV{row_idx}="B",Sheet3!$F$48,IF(AV{row_idx}="T",Sheet3!$F$49)))))'

                # Kemampuan Numerik 5.1 (kolom 61)
                numerik_cell = sheet.cell(row=row_idx, column=61)
                numerik_cell.value = f'=IF(AW{row_idx}="R",Sheet3!$F$40,IF(AW{row_idx}="K",Sheet3!$F$41,IF(AW{row_idx}="C",Sheet3!$F$42,IF(AW{row_idx}="B",Sheet3!$F$43,IF(AW{row_idx}="T",Sheet3!$F$44)))))'

                # Sistematika Kerja/ C D R.1 (kolom 62)
                sistematika_cell = sheet.cell(row=row_idx, column=62)
                sistematika_cell.value = f'=IF(AX{row_idx}="R",Sheet3!$F$60,IF(AX{row_idx}="K",Sheet3!$F$61,IF(AX{row_idx}="C",Sheet3!$F$62,IF(AX{row_idx}="B",Sheet3!$F$63,IF(AX{row_idx}="T",Sheet3!$F$64)))))'

                # Orientasi Hasil/ N G.1 (kolom 63)
                orientasi_cell = sheet.cell(row=row_idx, column=63)
                orientasi_cell.value = f'=IF(AY{row_idx}="R",Sheet3!$F$50,IF(AY{row_idx}="K",Sheet3!$F$51,IF(AY{row_idx}="C",Sheet3!$F$52,IF(AY{row_idx}="B",Sheet3!$F$53,IF(AY{row_idx}="T",Sheet3!$F$54)))))'

                # Fleksibilitas/ T V.1 (kolom 64)
                fleksibilitas_cell = sheet.cell(row=row_idx, column=64)
                fleksibilitas_cell.value = f'=IF(AZ{row_idx}="R",Sheet3!$F$55,IF(AZ{row_idx}="K",Sheet3!$F$56,IF(AZ{row_idx}="C",Sheet3!$F$57,IF(AZ{row_idx}="B",Sheet3!$F$58,IF(AZ{row_idx}="T",Sheet3!$F$59)))))'

                # Motivasi Berprestasi/ A.1 (kolom 65)
                motivasi_cell = sheet.cell(row=row_idx, column=65)
                motivasi_cell.value = f'=IF(BA{row_idx}="R",Sheet3!$F$65,IF(BA{row_idx}="K",Sheet3!$F$66,IF(BA{row_idx}="C",Sheet3!$F$67,IF(BA{row_idx}="B",Sheet3!$F$68,IF(BA{row_idx}="T",Sheet3!$F$69)))))'

                # Kerjasama/ P I.1 (kolom 66)
                kerjasama_cell = sheet.cell(row=row_idx, column=66)
                kerjasama_cell.value = f'=IF(BB{row_idx}="R",Sheet3!$F$70,IF(BB{row_idx}="K",Sheet3!$F$71,IF(BB{row_idx}="C",Sheet3!$F$72,IF(BB{row_idx}="B",Sheet3!$F$73,IF(BB{row_idx}="T",Sheet3!$F$74)))))'

                # Keterampilan Interpersonal/ B S.1 (kolom 67)
                interpersonal_cell = sheet.cell(row=row_idx, column=67)
                interpersonal_cell.value = f'=IF(BC{row_idx}="R",Sheet3!$F$75,IF(BC{row_idx}="K",Sheet3!$F$76,IF(BC{row_idx}="C",Sheet3!$F$77,IF(BC{row_idx}="B",Sheet3!$F$78,IF(BC{row_idx}="T",Sheet3!$F$79)))))'

                # Stabilitas Emosi/ E PHQ.1 (kolom 68)
                emosi_cell = sheet.cell(row=row_idx, column=68)
                emosi_cell.value = f'=IF(BD{row_idx}="R",Sheet3!$F$80,IF(BD{row_idx}="K",Sheet3!$F$81,IF(BD{row_idx}="C",Sheet3!$F$82,IF(BD{row_idx}="B",Sheet3!$F$83,IF(BD{row_idx}="T",Sheet3!$F$84)))))'

                # Pegembangan Diri/ W.1 (kolom 69)
                pengembangan_cell = sheet.cell(row=row_idx, column=69)
                pengembangan_cell.value = f'=IF(BE{row_idx}="R",Sheet3!$F$85,IF(BE{row_idx}="K",Sheet3!$F$86,IF(BE{row_idx}="C",Sheet3!$F$87,IF(BE{row_idx}="B",Sheet3!$F$88,IF(BE{row_idx}="T",Sheet3!$F$89)))))'

                # Mengelola Perubahan/ Z K.1 (kolom 70)
                perubahan_cell = sheet.cell(row=row_idx, column=70)
                perubahan_cell.value = f'=IF(BF{row_idx}="R",Sheet3!$J$90,IF(BF{row_idx}="K",Sheet3!$F$91,IF(BF{row_idx}="C",Sheet3!$F$92,IF(BF{row_idx}="B",Sheet3!$F$93,IF(BF{row_idx}="T",Sheet3!$F$94)))))'

            except Exception as e:
                print(f"Error setting .1 formulas: {e}")

            return True
        except Exception as e:
            print(f"Error applying psikogram formulas: {e}")
            return False

    def set_formulas_direct(self, sheet, header_row):
        """
        Fungsi untuk mengatur formula Excel secara langsung pada sheet
        """
        try:
            # Looping untuk setiap baris data (di bawah header)
            for row_idx in range(header_row + 1, sheet.max_row + 1):
                # Formula untuk SDR/SDRI
                sdr_sdri_cell = sheet.cell(row=row_idx, column=self.get_excel_column_index("SDR/SDRI"))
                sdr_sdri_cell.value = f'=IF(F{row_idx}="P","Sdri.","Sdr.")'
                
                # Formula untuk Keterangan PHQ
                ket_phq_cell = sheet.cell(row=row_idx, column=self.get_excel_column_index("Keterangan PHQ"))
                ket_phq_cell.value = f'=IFS(I{row_idx}<5,"Tidak ada",I{row_idx}<10,"Ringan",I{row_idx}<15,"Sedang",I{row_idx}<20,"Cukup Berat",I{row_idx}<28,"Parah")'
                
                # Formula untuk IQ
                iq_cell = sheet.cell(row=row_idx, column=self.get_excel_column_index("IQ "))
                iq_cell.value = f'=SUM(L{row_idx}:P{row_idx})/5'
                
                # Formula untuk Unnamed: 16
                unnamed_16_cell = sheet.cell(row=row_idx, column=self.get_excel_column_index("Unnamed: 16"))
                unnamed_16_cell.value = f'=(M{row_idx}+O{row_idx})/2'
                
                # Formula untuk KLASIFIKASI
                klasifikasi_cell = sheet.cell(row=row_idx, column=self.get_excel_column_index("KLASIFIKASI"))
                klasifikasi_cell.value = f'=IFS(K{row_idx}<79,"Rendah",K{row_idx}<90,"Dibawah Rata-Rata",K{row_idx}<110,"Rata-Rata",K{row_idx}<120,"Diatas Rata-Rata",K{row_idx}>119,"Superior")'
                
                # Formula untuk C (Coding)
                c_coding_cell = sheet.cell(row=row_idx, column=self.get_excel_column_index("C (Coding)"))
                c_coding_cell.value = f'=IFS(AE{row_idx}=1,9,AE{row_idx}=2,8,AE{row_idx}=3,7,AE{row_idx}=4,6,AE{row_idx}=5,5,AE{row_idx}=6,4,AE{row_idx}=7,3,AE{row_idx}=8,2,AE{row_idx}=9,1)'
                
                # Formula untuk NG
                ng_cell = sheet.cell(row=row_idx, column=self.get_excel_column_index("NG"))
                ng_cell.value = f'=(S{row_idx}+T{row_idx})/2'
                
                # Formula untuk CDR
                cdr_cell = sheet.cell(row=row_idx, column=self.get_excel_column_index("CDR"))
                cdr_cell.value = f'=(AE{row_idx}+AG{row_idx}+AH{row_idx})/3'
                
                # Formula untuk TV
                tv_cell = sheet.cell(row=row_idx, column=self.get_excel_column_index("TV"))
                tv_cell.value = f'=(Y{row_idx}+Z{row_idx})/2'
                
                # Formula untuk PI
                pi_cell = sheet.cell(row=row_idx, column=self.get_excel_column_index("PI"))
                pi_cell.value = f'=(W{row_idx}+X{row_idx})/2'
                
                # Formula untuk BS
                bs_cell = sheet.cell(row=row_idx, column=self.get_excel_column_index("BS"))
                bs_cell.value = f'=(AA{row_idx}+AB{row_idx})/2'
                
                # Formula untuk ZK
                zk_cell = sheet.cell(row=row_idx, column=self.get_excel_column_index("ZK"))
                zk_cell.value = f'=(AI{row_idx}+AK{row_idx})/2'

                # Terapkan rumus psikogram
                self.apply_psikogram_formulas(sheet, row_idx)

            print("Formula berhasil diaplikasikan langsung ke sheet Excel")
            return True
        except Exception as e:
            print(f"Error saat menerapkan formula: {e}")
            return False

    def get_w_column_index(self):
        """Helper function untuk mendapatkan indeks kolom W yang benar"""
        return 38  # Indeks kolom W di Excel

    def get_w_value(self, row):
        """Helper function untuk mengambil nilai W dari baris tertentu"""
        w_col = self.get_w_column_index()
        value = self.get_cell_text(row, w_col)
        print(f"DEBUG - Mengambil nilai W dari baris {row}, kolom {w_col}: '{value}'")
        return value

    def set_w_value(self, row, value):
        """Helper function untuk menyimpan nilai W ke baris tertentu"""
        w_col = self.get_w_column_index()
        self.table.setItem(row, w_col, QTableWidgetItem(str(value)))
        print(f"DEBUG - Menyimpan nilai W ke baris {row}, kolom {w_col}: '{value}'")

    def get_sheet3_reference(self, column_name, value):
        """
        Fungsi untuk mendapatkan referensi nilai dari Sheet3 berdasarkan kolom dan nilai
        """
        # Mapping untuk setiap kolom dan nilai referensinya di Sheet3
        sheet3_references = {
            "Logika Berpikir 1": {
                "R": "menunjukkan kesulitan dalam mengidentifikasi pola logis, yang mengindikasikan kemampuan penalaran induktif yang terbatas.",
                "K": "menunjukkan kesulitan dalam mempertahankan konsistensi logis dalam argumennya, terkadang membuat kesimpulan yang tidak sepenuhnya didukung oleh bukti.",
                "C": "mampu menerapkan prinsip-prinsip logika dasar dalam pemecahan masalah, meskipun terkadang membutuhkan waktu untuk mencapai kesimpulan yang tepat.",
                "B": "menunjukkan kemampuan yang kuat dalam mengidentifikasi dan menerapkan pola logis dalam berbagai konteks, menunjukkan kemampuan penalaran yang tinggi.",
                "T": "menunjukkan kemampuan luar biasa dalam menerapkan prinsip-prinsip logika tingkat lanjut, mampu memecahkan masalah yang kompleks dengan mudah."
            },
            "Daya Analisa 3": {
                "R": "Ybs menunjukkan kesulitan dalam mengidentifikasi komponen-komponen penting dari suatu masalah, sehingga pemahaman terhadap hubungan sebab-akibat menjadi terbatas.",
                "K": "Ybs menunjukkan keterbatasan dalam mengidentifikasi pola-pola yang mendasari suatu permasalahan, sehingga pemahaman terhadap dinamika masalah kurang mendalam.",
                "C": "Ybs mampu mengidentifikasi komponen-komponen utama dari suatu permasalahan dan memahami hubungan dasar antara faktor-faktor tersebut.",
                "B": "Ybs menunjukkan kemampuan yang baik dalam mengidentifikasi pola-pola yang kompleks dan memahami hubungan sebab-akibat yang mendalam.",
                "T": "Ybs menunjukkan kemampuan yang sangat tinggi dalam mengidentifikasi pola-pola yang abstrak dan memahami hubungan sebab-akibat yang sangat kompleks."
            },
            "Kemampuan Verbal 2 dam 4": {
                "R": "Kemudian, menunjukkan kesulitan dalam memahami instruksi verbal yang kompleks, seringkali memerlukan penjelasan yang berulang.",
                "K": "Kemudian, menunjukkan pemahaman yang kurang memadai terhadap nuansa makna dalam bahasa, yang mempengaruhi kemampuan mereka dalam berkomunikasi secara efektif.",
                "C": "Kemudian, memiliki pemahaman yang cukup baik terhadap kosa kata umum dan mampu mengikuti instruksi verbal dengan cukup baik.",
                "B": "Kemudian, menunjukkan pemahaman yang baik terhadap kosa kata yang luas dan mampu memahami nuansa makna dalam bahasa.",
                "T": "Kemudian, memiliki penguasaan yang sangat baik terhadap kosa kata yang luas dan kompleks, mampu memahami dan menggunakan bahasa dengan sangat presisi."
            },
            "Kemampuan Numerik 5": {
                "R": "Selain itu, menunjukkan kesulitan yang signifikan dalam memahami konsep dasar matematika seperti persentase dan rasio.",
                "K": "Selain itu, menunjukkan pemahaman yang kurang memadai terhadap konsep matematika dasar, yang mempengaruhi kemampuan mereka dalam menyelesaikan masalah numerik.",
                "C": "Selain itu, memiliki pemahaman yang cukup baik terhadap konsep matematika dasar dan mampu melakukan perhitungan aritmatika dengan cukup akurat.",
                "B": "Selain itu, menunjukkan pemahaman yang baik terhadap konsep matematika dan mampu melakukan perhitungan aritmatika dengan cepat dan akurat.",
                "T": "Selain itu, memiliki pemahaman yang mendalam terhadap konsep matematika tingkat lanjut dan mampu melakukan perhitungan aritmatika yang kompleks dengan sangat cepat dan akurat."
            },
            "Sistematika Kerja/ C D R": {
                "R": "Kemudian, menunjukkan kesulitan dalam membuat rencana kerja yang terstruktur, seringkali memulai tugas tanpa persiapan yang memadai.",
                "K": "Kemudian, menunjukkan upaya yang kurang maksimal dalam merencanakan tugas-tugas, terkadang mengabaikan detail-detail penting dalam proses perencanaan.",
                "C": "Kemudian, mampu membuat rencana kerja yang cukup terstruktur, meskipun terkadang membutuhkan pengawasan tambahan.",
                "B": "Kemudian, menunjukkan kemampuan yang baik dalam membuat rencana kerja yang terstruktur dan detail, mampu mengantisipasi potensi hambatan.",
                "T": "Kemudian, memiliki kemampuan yang sangat tinggi dalam membuat rencana kerja yang komprehensif dan strategis, mampu mengintegrasikan berbagai aspek pekerjaan."
            },
            "Orientasi Hasil/ N G": {
                "R": "menunjukkan kesulitan dalam mempertahankan fokus pada penyelesaian tugas, seringkali teralihkan oleh gangguan eksternal.",
                "K": "menunjukkan upaya yang kurang maksimal dalam menyelesaikan tugas, terkadang mengabaikan detail-detail penting yang mempengaruhi hasil akhir.",
                "C": "mampu menyelesaikan tugas-tugas yang diberikan dengan cukup baik, meskipun terkadang membutuhkan pengawasan tambahan.",
                "B": "menunjukkan komitmen yang kuat dalam menyelesaikan tugas-tugas yang diberikan, selalu berusaha untuk mencapai hasil yang terbaik.",
                "T": "memiliki dorongan yang sangat kuat untuk mencapai hasil yang luar biasa, selalu melampaui ekspektasi dalam menyelesaikan tugas-tugas yang diberikan."
            },
            "Fleksibilitas/ T V": {
                "R": "Ybs menunjukkan kesulitan dalam beradaptasi dengan perubahan situasi, cenderung kaku dalam menghadapi hal-hal baru.",
                "K": "Ybs menunjukkan upaya yang kurang maksimal dalam menyesuaikan diri dengan perubahan, terkadang merasa tidak nyaman dengan hal-hal baru.",
                "C": "Ybs mampu menyesuaikan diri dengan perubahan situasi dengan cukup baik, meskipun terkadang membutuhkan waktu untuk beradaptasi",
                "B": "Ybs menunjukkan kemampuan yang baik dalam beradaptasi dengan perubahan situasi, mampu dengan cepat menyesuaikan diri dengan hal-hal baru.",
                "T": "Ybs memiliki kemampuan yang sangat tinggi dalam beradaptasi dengan perubahan situasi, mampu dengan cepat dan efektif menyesuaikan diri dengan hal-hal baru."
            },
            "Motivasi Berprestasi/ A": {
                "R": "menunjukkan kurangnya dorongan untuk mencapai target yang ditetapkan, seringkali merasa puas dengan hasil yang minimal.",
                "K": "menunjukkan upaya yang kurang maksimal dalam mencapai target, terkadang kehilangan fokus dan motivasi di tengah jalan.",
                "C": "mampu mencapai target yang ditetapkan dengan cukup baik, meskipun terkadang membutuhkan dorongan tambahan.",
                "B": "menunjukkan dorongan yang kuat untuk mencapai target yang ditetapkan, selalu berusaha untuk melampaui ekspektasi.",
                "T": "memiliki dorongan yang sangat kuat untuk mencapai prestasi yang luar biasa, selalu menetapkan target yang sangat tinggi dan menantang"
            },
            "Kerjasama/ P I": {
                "R": "Tak hanya itu, ybs juga menunjukkan kesulitan dalam berpartisipasi dalam kerja kelompok, seringkali lebih memilih untuk bekerja sendiri.",
                "K": "Tak hanya itu, ybs juga menunjukkan upaya yang kurang maksimal dalam berkontribusi pada kerja kelompok, terkadang mengabaikan tanggung jawab yang diberikan.",
                "C": "Tak hanya itu, ybs juga mampu berpartisipasi dalam kerja kelompok dengan cukup baik, meskipun terkadang membutuhkan arahan tambahan.",
                "B": "Tak hanya itu, ybs juga menunjukkan kemampuan yang baik dalam berkolaborasi dengan anggota tim, selalu berusaha untuk memberikan kontribusi yang positif.",
                "T": "Tak hanya itu, ybs juga memiliki kemampuan yang sangat tinggi dalam membangun dan memimpin tim yang efektif, mampu menginspirasi dan memotivasi anggota tim untuk mencapai tujuan bersama."
            },
            "Keterampilan Interpersonal/ B S": {
                "R": "Ybs menunjukkan kesulitan dalam memahami dan merespons emosi orang lain, yang menyebabkan seringnya terjadi kesalahpahaman dalam interaksi sosial.",
                "K": "Ybs menunjukkan upaya yang kurang maksimal dalam membangun hubungan sosial yang efektif, terkadang mengalami kesulitan dalam berkomunikasi secara asertif.",
                "C": "Ybs mampu membangun dan mempertahankan hubungan sosial yang cukup baik, meskipun terkadang membutuhkan arahan tambahan.",
                "B": "Ybs menunjukkan kemampuan yang baik dalam membangun dan mempertahankan hubungan sosial yang positif, mampu menciptakan suasana yang nyaman dan mendukung.",
                "T": "Ybs memiliki kemampuan yang sangat tinggi dalam membangun dan memimpin hubungan sosial yang kompleks, mampu menginspirasi dan memotivasi orang lain."
            },
            "Stabilitas Emosi/ E PHQ": {
                "R": "Selain itu juga, menunjukkan fluktuasi emosi yang signifikan, seringkali bereaksi berlebihan terhadap situasi yang menekan.",
                "K": "Selain itu juga, menunjukkan upaya yang kurang maksimal dalam mengendalikan emosi, terkadang mengalami kesulitan dalam menghadapi situasi yang sulit.",
                "C": "Selain itu juga, mampu mengendalikan emosi dengan cukup baik dalam situasi yang relatif stabil, meskipun terkadang membutuhkan waktu untuk beradaptasi dengan perubahan.",
                "B": "Selain itu juga, menunjukkan kemampuan yang baik dalam mengendalikan emosi, mampu menjaga ketenangan dalam situasi yang menekan.",
                "T": "Selain itu juga, memiliki kemampuan yang sangat tinggi dalam mengendalikan emosi, mampu menjaga keseimbangan emosional dalam situasi yang sangat menekan."
            },
            "Pegembangan Diri/ W": {
                "R": "menunjukkan kurangnya inisiatif dalam mencari peluang untuk meningkatkan pengetahuan dan keterampilan diri.",
                "K": "menunjukkan upaya yang kurang maksimal dalam mengembangkan diri, terkadang kurang konsisten dalam mengikuti program pelatihan atau pengembangan",
                "C": "mampu mengikuti program pelatihan atau pengembangan diri dengan cukup baik, meskipun terkadang membutuhkan dorongan tambahan.",
                "B": "menunjukkan inisiatif yang kuat dalam mencari peluang untuk mengembangkan diri, selalu berusaha untuk meningkatkan pengetahuan dan keterampilan.",
                "T": "memiliki dorongan yang sangat kuat untuk mengembangkan diri, selalu mencari tantangan dan peluang untuk pertumbuhan pribadi."
            },
            "Mengelola Perubahan/ Z K": {
                "R": "serta menunjukkan resistensi yang kuat terhadap perubahan, seringkali merasa tidak nyaman dan tertekan dalam situasi baru.",
                "K": "serta untuk mengubah pola pikir dan perilaku dalam menghadapi perubahan masih berkembang, yang menyebabkan terkadang terjadi kesulitan dalam beradaptasi.",
                "C": "serta mampu menyesuaikan diri dengan perubahan situasi dengan cukup baik, meskipun terkadang membutuhkan waktu untuk beradaptasi.",
                "B": "serta menunjukkan kemampuan yang baik dalam beradaptasi dengan perubahan, mampu dengan cepat menyesuaikan diri dengan situasi baru.",
                "T": "serta memiliki kemampuan yang sangat tinggi dalam mengelola perubahan, mampu memimpin dan menginspirasi orang lain untuk beradaptasi dengan situasi baru."
            }
        }

        # Dapatkan referensi dari mapping
        if column_name in sheet3_references and value in sheet3_references[column_name]:
            return sheet3_references[column_name][value]
        return ""  # Return string kosong jika tidak ada referensi yang cocok

if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = ExcelViewerApp()
    window.show()
    sys.exit(app.exec_())
